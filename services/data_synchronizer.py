"""
Модуль для синхронизации данных между маркетплейсами
"""
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from typing import Dict, List, Tuple, Optional
from utils.excel_reader import ExcelReader
from utils.excel_writer import ExcelWriter
from config.config import FILE_CONFIGS, is_excluded_column
from services.ai_comparator import AIComparator
from utils.logger_config import setup_logger
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

# Создаем глобальный логгер
logger = setup_logger('data_sync')  

class DimensionsSynchronizer:
    """Синхронизация композитных габаритов (Длина/Ширина/Высота)"""

    # Маппинг столбцов для каждого маркетплейса
    DIMENSIONS_MAPPING = {
        'wildberries': {
            'length': 'Длина упаковки (целое число)',
            'width': 'Ширина упаковки (целое число)',
            'height': 'Высота упаковки (целое число)',
            'unit': 'cm'
        },
        'ozon': {
            'length': 'Длина упаковки, мм*',
            'width': 'Ширина упаковки, мм*',
            'height': 'Высота упаковки, мм*',
            'unit': 'mm'
        },
        'yandex': {
            'composite': 'Габариты с упаковкой, см',
            'unit': 'cm'
        }
    }

    @staticmethod
    def parse_composite_dimensions(value: str) -> Optional[Dict[str, float]]:
        """
        Парсит строку "71/68/197" в словарь {length, width, height}
        """
        if pd.isna(value) or not str(value).strip():
            return None

        try:
            parts = str(value).strip().split('/')
            if len(parts) != 3:
                return None

            dimensions = {
                'length': float(parts[0].strip()),
                'width': float(parts[1].strip()),
                'height': float(parts[2].strip())
            }

            if all(v > 0 for v in dimensions.values()):
                return dimensions

        except (ValueError, AttributeError):
            pass

        return None

    @staticmethod
    def format_composite_dimensions(length: float, width: float, height: float) -> str:
        """
        Форматирует габариты в строку "Длина/Ширина/Высота"
        """
        def smart_format(val):
            if abs(val - round(val)) < 0.01:
                return str(int(round(val)))
            return f"{val:.1f}"

        return f"{smart_format(length)}/{smart_format(width)}/{smart_format(height)}"

    @staticmethod
    def mm_to_cm(value: float) -> float:
        """Конвертирует миллиметры в сантиметры"""
        return value / 10

    @staticmethod
    def cm_to_mm(value: float) -> float:
        """Конвертирует сантиметры в миллиметры"""
        return value * 10

    @classmethod
    def sync_dimensions(cls, dfs: Dict[str, pd.DataFrame]) -> int:
        """
        Синхронизирует габариты между всеми маркетплейсами
        """
        logger.info("=" * 80)
        logger.info("🔧 НАЧАЛО СИНХРОНИЗАЦИИ ГАБАРИТОВ (DimensionsSynchronizer)")
        logger.info("=" * 80)
        
        synced_count = 0

        # Получаем названия столбцов артикулов
        article_cols = {
            'wildberries': 'Артикул продавца',
            'ozon': 'Артикул*',
            'yandex': 'Ваш SKU *'
        }

        # Создаём маппинг артикул → данные
        yandex_dimensions = {}
        wb_dimensions = {}
        ozon_dimensions = {}

        # ==================== ЭТАП 1: ЧТЕНИЕ ДАННЫХ ИЗ ЯНДЕКС ====================
        logger.info("\n📖 ЭТАП 1: Чтение габаритов из Яндекс (композитный формат)")
        
        yandex_col = cls.DIMENSIONS_MAPPING['yandex']['composite']
        logger.info(f"   Ищем столбец: '{yandex_col}'")
        
        if 'yandex' not in dfs:
            logger.warning("   ❌ DataFrame 'yandex' отсутствует!")
        elif yandex_col not in dfs['yandex'].columns:
            logger.warning(f"   ❌ Столбец '{yandex_col}' не найден в Яндекс!")
            logger.info(f"   Доступные столбцы: {list(dfs['yandex'].columns)[:10]}...")
        else:
            logger.info(f"   ✅ Столбец '{yandex_col}' найден")
            
            rows_checked = 0
            rows_with_article = 0
            rows_with_dimensions = 0
            
            for idx, row in dfs['yandex'].iterrows():
                rows_checked += 1
                article = row.get(article_cols['yandex'])
                
                if pd.notna(article) and str(article).strip():
                    rows_with_article += 1
                    composite = row.get(yandex_col)
                    
                    if rows_with_article <= 5:
                        logger.debug(f"   [{rows_with_article}] Артикул: '{article}', Габариты: '{composite}'")
                    
                    dimensions = cls.parse_composite_dimensions(composite)
                    if dimensions:
                        rows_with_dimensions += 1
                        yandex_dimensions[str(article).strip()] = dimensions
                        if rows_with_dimensions <= 3:
                            logger.info(f"   ✓ Яндекс [{article}]: {dimensions['length']}/{dimensions['width']}/{dimensions['height']} см")
            
            logger.info(f"   📊 Итого Яндекс: проверено {rows_checked} строк, с артикулами {rows_with_article}, с габаритами {rows_with_dimensions}")

        # ==================== ЭТАП 2: ЧТЕНИЕ ДАННЫХ ИЗ WB ====================
        logger.info("\n📖 ЭТАП 2: Чтение габаритов из WB (раздельные столбцы, см)")
        
        if 'wildberries' not in dfs:
            logger.warning("   ❌ DataFrame 'wildberries' отсутствует!")
        else:
            wb_map = cls.DIMENSIONS_MAPPING['wildberries']
            df_wb = dfs['wildberries']
            
            missing_cols = []
            for col in [wb_map['length'], wb_map['width'], wb_map['height']]:
                if col not in df_wb.columns:
                    missing_cols.append(col)
                    logger.warning(f"   ❌ Столбец '{col}' не найден!")
                else:
                    logger.info(f"   ✅ Столбец '{col}' найден")
            
            if missing_cols:
                logger.error(f"   ⛔ Пропускаю WB из-за отсутствия столбцов: {missing_cols}")
            else:
                rows_with_dimensions = 0
                
                for idx, row in df_wb.iterrows():
                    article = row.get(article_cols['wildberries'])
                    
                    if pd.notna(article) and str(article).strip():
                        article_str = str(article).strip()
                        length = row.get(wb_map['length'])
                        width = row.get(wb_map['width'])
                        height = row.get(wb_map['height'])

                        if all(pd.notna(v) and str(v).strip() for v in [length, width, height]):
                            try:
                                wb_dimensions[article_str] = {
                                    'length': float(length),
                                    'width': float(width),
                                    'height': float(height)
                                }
                                rows_with_dimensions += 1
                                if rows_with_dimensions <= 3:
                                    logger.info(f"   ✓ WB [{article_str}]: {length}/{width}/{height} см")
                            except ValueError:
                                pass
                
                logger.info(f"   📊 Итого WB: {rows_with_dimensions} артикулов с полными габаритами")

        # ==================== ЭТАП 3: ЧТЕНИЕ ДАННЫХ ИЗ OZON ====================
        logger.info("\n📖 ЭТАП 3: Чтение габаритов из Ozon (раздельные столбцы, мм)")
        
        if 'ozon' not in dfs:
            logger.warning("   ❌ DataFrame 'ozon' отсутствует!")
        else:
            ozon_map = cls.DIMENSIONS_MAPPING['ozon']
            df_ozon = dfs['ozon']
            
            missing_cols = []
            for col in [ozon_map['length'], ozon_map['width'], ozon_map['height']]:
                if col not in df_ozon.columns:
                    missing_cols.append(col)
                    logger.warning(f"   ❌ Столбец '{col}' не найден!")
                else:
                    logger.info(f"   ✅ Столбец '{col}' найден")
            
            if missing_cols:
                logger.error(f"   ⛔ Пропускаю Ozon из-за отсутствия столбцов: {missing_cols}")
            else:
                rows_with_dimensions = 0
                
                for idx, row in df_ozon.iterrows():
                    article = row.get(article_cols['ozon'])
                    
                    if pd.notna(article) and str(article).strip():
                        article_str = str(article).strip()
                        length_mm = row.get(ozon_map['length'])
                        width_mm = row.get(ozon_map['width'])
                        height_mm = row.get(ozon_map['height'])

                        if all(pd.notna(v) and str(v).strip() for v in [length_mm, width_mm, height_mm]):
                            try:
                                ozon_dimensions[article_str] = {
                                    'length': cls.mm_to_cm(float(length_mm)),
                                    'width': cls.mm_to_cm(float(width_mm)),
                                    'height': cls.mm_to_cm(float(height_mm))
                                }
                                rows_with_dimensions += 1
                                if rows_with_dimensions <= 3:
                                    logger.info(f"   ✓ Ozon [{article_str}]: {length_mm}/{width_mm}/{height_mm} мм → {ozon_dimensions[article_str]['length']}/{ozon_dimensions[article_str]['width']}/{ozon_dimensions[article_str]['height']} см")
                            except ValueError:
                                pass
                
                logger.info(f"   📊 Итого Ozon: {rows_with_dimensions} артикулов с полными габаритами")

        # ==================== СВОДКА СОБРАННЫХ ДАННЫХ ====================
        logger.info("\n📊 СВОДКА СОБРАННЫХ ДАННЫХ:")
        logger.info(f"   • Яндекс: {len(yandex_dimensions)} артикулов с габаритами")
        logger.info(f"   • WB: {len(wb_dimensions)} артикулов с габаритами")
        logger.info(f"   • Ozon: {len(ozon_dimensions)} артикулов с габаритами")

        # ==================== ЭТАП 4: СИНХРОНИЗАЦИЯ Яндекс → WB/Ozon ====================
        logger.info("\n🔄 ЭТАП 4: Синхронизация Яндекс → WB/Ozon")
        
        yandex_to_wb_count = 0
        yandex_to_ozon_count = 0
        
        for article, dimensions in yandex_dimensions.items():
            # Синхронизация в WB
            if 'wildberries' in dfs:
                df_wb = dfs['wildberries']
                wb_map = cls.DIMENSIONS_MAPPING['wildberries']

                mask = df_wb[article_cols['wildberries']].astype(str).str.strip() == article
                
                if mask.any():
                    idx = df_wb[mask].index[0]

                    for dim_key, col_name in [('length', wb_map['length']), ('width', wb_map['width']), ('height', wb_map['height'])]:
                        current_val = df_wb.at[idx, col_name]
                        if pd.isna(current_val) or not str(current_val).strip():
                            df_wb.at[idx, col_name] = dimensions[dim_key]
                            synced_count += 1
                            yandex_to_wb_count += 1
                            logger.debug(f"   [Яндекс→WB] {article}: {dim_key}={dimensions[dim_key]}")

            # Синхронизация в Ozon
            if 'ozon' in dfs:
                df_ozon = dfs['ozon']
                ozon_map = cls.DIMENSIONS_MAPPING['ozon']

                mask = df_ozon[article_cols['ozon']].astype(str).str.strip() == article
                
                if mask.any():
                    idx = df_ozon[mask].index[0]

                    for dim_key, col_name in [('length', ozon_map['length']), ('width', ozon_map['width']), ('height', ozon_map['height'])]:
                        current_val = df_ozon.at[idx, col_name]
                        if pd.isna(current_val) or not str(current_val).strip():
                            value_mm = int(cls.cm_to_mm(dimensions[dim_key]))
                            df_ozon.at[idx, col_name] = value_mm
                            synced_count += 1
                            yandex_to_ozon_count += 1
                            logger.debug(f"   [Яндекс→Ozon] {article}: {dim_key}={value_mm} мм")
        
        logger.info(f"   📊 Яндекс → WB: {yandex_to_wb_count} значений")
        logger.info(f"   📊 Яндекс → Ozon: {yandex_to_ozon_count} значений")

        # ==================== ЭТАП 5: СИНХРОНИЗАЦИЯ WB → Яндекс и Ozon ====================
        logger.info("\n🔄 ЭТАП 5: Синхронизация WB → Яндекс и Ozon")

        wb_to_yandex_count = 0
        wb_to_ozon_count = 0
        wb_skipped_yandex = 0

        for article, dimensions in wb_dimensions.items():
            # ============ В Яндекс ============
            if article in yandex_dimensions:
                wb_skipped_yandex += 1
                logger.debug(f"   [WB→Яндекс] ПРОПУСК {article}: уже есть данные в Яндексе")
            else:
                if 'yandex' in dfs:
                    df_yandex = dfs['yandex']
                    yandex_col = cls.DIMENSIONS_MAPPING['yandex']['composite']

                    mask = df_yandex[article_cols['yandex']].astype(str).str.strip() == article
                    
                    if mask.any():
                        idx = df_yandex[mask].index[0]
                        current_val = df_yandex.at[idx, yandex_col]
                        
                        if pd.isna(current_val) or not str(current_val).strip():
                            composite = cls.format_composite_dimensions(
                                dimensions['length'],
                                dimensions['width'],
                                dimensions['height']
                            )
                            df_yandex.at[idx, yandex_col] = composite
                            synced_count += 1
                            wb_to_yandex_count += 1
                            logger.info(f"   ✓ [WB→Яндекс] {article}: {composite}")

            # ============ В Ozon ============
            if 'ozon' in dfs:
                df_ozon = dfs['ozon']
                ozon_map = cls.DIMENSIONS_MAPPING['ozon']

                mask = df_ozon[article_cols['ozon']].astype(str).str.strip() == article
                
                if mask.any():
                    idx = df_ozon[mask].index[0]

                    for dim_key, col_name in [('length', ozon_map['length']), ('width', ozon_map['width']), ('height', ozon_map['height'])]:
                        current_val = df_ozon.at[idx, col_name]
                        if pd.isna(current_val) or not str(current_val).strip():
                            value_mm = int(cls.cm_to_mm(dimensions[dim_key]))
                            df_ozon.at[idx, col_name] = value_mm
                            synced_count += 1
                            wb_to_ozon_count += 1
                            logger.info(f"   ✓ [WB→Ozon] {article}: {dim_key}={value_mm} мм")

        logger.info(f"   📊 WB → Яндекс: {wb_to_yandex_count} значений записано")
        logger.info(f"   📊 WB → Яндекс: {wb_skipped_yandex} пропущено (уже есть в Яндексе)")
        logger.info(f"   📊 WB → Ozon: {wb_to_ozon_count} значений")

        # ==================== ЭТАП 6: СИНХРОНИЗАЦИЯ Ozon → Яндекс и WB ====================
        logger.info("\n🔄 ЭТАП 6: Синхронизация Ozon → Яндекс и WB")
        
        ozon_to_yandex_count = 0
        ozon_to_wb_count = 0
        ozon_skipped_yandex = 0
        
        for article, dimensions in ozon_dimensions.items():
            # В Яндекс
            if 'yandex' in dfs:
                df_yandex = dfs['yandex']
                yandex_col = cls.DIMENSIONS_MAPPING['yandex']['composite']

                mask = df_yandex[article_cols['yandex']].astype(str).str.strip() == article
                
                if mask.any():
                    idx = df_yandex[mask].index[0]
                    current_val = df_yandex.at[idx, yandex_col]
                    
                    if pd.isna(current_val) or not str(current_val).strip():
                        composite = cls.format_composite_dimensions(
                            dimensions['length'],
                            dimensions['width'],
                            dimensions['height']
                        )
                        df_yandex.at[idx, yandex_col] = composite
                        synced_count += 1
                        ozon_to_yandex_count += 1
                        logger.info(f"   ✓ [Ozon→Яндекс] {article}: {composite}")
                    else:
                        ozon_skipped_yandex += 1

            # В WB
            if 'wildberries' in dfs:
                df_wb = dfs['wildberries']
                wb_map = cls.DIMENSIONS_MAPPING['wildberries']

                mask = df_wb[article_cols['wildberries']].astype(str).str.strip() == article
                
                if mask.any():
                    idx = df_wb[mask].index[0]

                    for dim_key, col_name in [('length', wb_map['length']), ('width', wb_map['width']), ('height', wb_map['height'])]:
                        current_val = df_wb.at[idx, col_name]
                        if pd.isna(current_val) or not str(current_val).strip():
                            df_wb.at[idx, col_name] = dimensions[dim_key]
                            synced_count += 1
                            ozon_to_wb_count += 1
                            logger.info(f"   ✓ [Ozon→WB] {article}: {dim_key}={dimensions[dim_key]}")
        
        logger.info(f"   📊 Ozon → Яндекс: {ozon_to_yandex_count} значений записано")
        logger.info(f"   📊 Ozon → Яндекс: {ozon_skipped_yandex} пропущено (ячейка заполнена)")
        logger.info(f"   📊 Ozon → WB: {ozon_to_wb_count} значений")

        # ==================== ИТОГОВАЯ СВОДКА ====================
        logger.info("\n" + "=" * 80)
        logger.info(f"✅ ГАБАРИТЫ: синхронизировано {synced_count} значений")
        logger.info("=" * 80)
        
        return synced_count


class DataSynchronizer:
    """Класс для синхронизации данных между тремя маркетплейсами"""
    
    def __init__(self, comparison_result: Dict, ai_comparator=None):
        self.comparison_result = comparison_result
        self.article_columns = {
            'wildberries': 'Артикул продавца',
            'ozon': 'Артикул*',
            'yandex': 'Ваш SKU *'
        }
        self.changes_log = {
            'wildberries': [],
            'ozon': [],
            'yandex': []
        }
        self.original_file_paths = {}
        self.ai_comparator = ai_comparator
        
        
        # ДОБАВЬТЕ: Кэш validation для каждого столбца
        self.column_validations = {}  # {marketplace: {column_name: [allowed_values]}}
        self.original_column_names = {}
        # ДОБАВЬТЕ ЭТУ СТРОКУ:
        self.ai_validation_log = []  # Логи AI-сопоставлений
        self.ai_cache = {}
        logger.info("Инициализация DataSynchronizer")
        logger.debug(f"AI comparator передан: {ai_comparator is not None}")
    
    def _align_articles(self, dfs: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
        """
        Выравнивает артикулы между маркетплейсами - добавляет отсутствующие строки
        Args:
            dfs: словарь с DataFrame для каждого маркетплейса
        Returns:
            Обновленные DataFrame с добавленными артикулами
        """
        logger.info("\n" + "="*60)
        logger.info("ВЫРАВНИВАНИЕ АРТИКУЛОВ МЕЖДУ МАРКЕТПЛЕЙСАМИ")
        logger.info("="*60)
        
        # Собираем все уникальные артикулы из всех маркетплейсов
        all_articles = set()
        for marketplace in ['wildberries', 'ozon', 'yandex']:
            article_col = self.article_columns[marketplace]
            if article_col in dfs[marketplace].columns:
                articles = dfs[marketplace][article_col].dropna().astype(str).str.strip()
                articles = articles[articles != '']  # Убираем пустые
                
                # Фильтрация: Убираем описания полей и слишком длинные строки
                articles = articles[
                    ~articles.str.contains(
                        'идентифицировать|описание|заполнить|пример|название товара|по которому',
                        case=False,
                        na=False
                    )
                ]
                # Убираем строки длиннее 50 символов (скорее всего описание)
                articles = articles[articles.str.len() < 50]
                all_articles.update(articles.tolist())
                logger.info(f"📊 {marketplace.upper()}: {len(articles)} артикулов")
        
        logger.info(f"\n🔍 Всего уникальных артикулов: {len(all_articles)}")
        
        # Для каждого маркетплейса проверяем недостающие артикулы
        total_added = 0
        for marketplace in ['wildberries', 'ozon', 'yandex']:
            article_col = self.article_columns[marketplace]
            if article_col not in dfs[marketplace].columns:
                logger.warning(f"⚠️ {marketplace.upper()}: столбец '{article_col}' не найден, пропускаю")
                continue
            
            # Сбрасываем индексы ПЕРЕД обработкой!
            dfs[marketplace] = dfs[marketplace].reset_index(drop=True)
            
            # Существующие артикулы
            df = dfs[marketplace]
            
            # 🆕 НОВЫЙ ПОДХОД: Находим строки с заполненными артикулами напрямую в DataFrame
            article_series = df[article_col].dropna().astype(str).str.strip()
            article_series = article_series[article_series != '']
            
            # Фильтрация
            valid_mask = (
                ~article_series.str.contains(
                    'идентифицировать|описание|заполнить|пример|название товара|по которому',
                    case=False,
                    na=False
                ) & 
                (article_series.str.len() < 50)
            )
            article_series = article_series[valid_mask]
            
            # 🆕 ИСПРАВЛЕНИЕ: Получаем ПОЗИЦИОННЫЙ индекс последней заполненной строки
            if len(article_series) > 0:
                # Получаем label индекс последней заполненной строки
                last_label_idx = article_series.index[-1]
                # Конвертируем в позиционный индекс
                last_filled_position = df.index.get_loc(last_label_idx)
            else:
                last_filled_position = -1
            
            existing_articles_set = set(article_series.tolist())
            
            # Находим недостающие
            missing_articles = all_articles - existing_articles_set
            
            if not missing_articles:
                logger.info(f"✅ {marketplace.upper()}: все артикулы присутствуют")
                continue
            
            logger.info(f"\n➕ {marketplace.upper()}: добавляю {len(missing_articles)} артикулов")
            
            # Создаем новые строки для недостающих артикулов
            new_rows = []
            for article in sorted(missing_articles):
                # Создаем пустую строку со всеми столбцами
                new_row = {col: None for col in df.columns}
                # Заполняем только артикул
                new_row[article_col] = article
                new_rows.append(new_row)
            
            # Вставляем новые строки СРАЗУ ПОСЛЕ последней заполненной!
            if new_rows:
                new_df = pd.DataFrame(new_rows)
                
                if last_filled_position >= 0:
                    # Есть заполненные строки - вставляем после них
                    # Используем ПОЗИЦИОННЫЙ индекс!
                    before = df.iloc[:last_filled_position + 1].copy()
                    after = df.iloc[last_filled_position + 1:].copy()
                    
                    # Склеиваем: заполненные + новые + пустые
                    dfs[marketplace] = pd.concat([before, new_df, after], ignore_index=True)
                    
                    logger.info(f" ✓ Добавлено {len(new_rows)} строк после позиции {last_filled_position}")
                else:
                    # Нет заполненных строк - добавляем в начало
                    dfs[marketplace] = pd.concat([new_df, df], ignore_index=True)
                    logger.info(f" ✓ Добавлено {len(new_rows)} строк в начало")
                
                total_added += len(new_rows)
                logger.info(f" 📊 Было: {len(df)}, стало: {len(dfs[marketplace])}")
        
        if total_added > 0:
            logger.info(f"\n✅ Итого добавлено {total_added} новых строк во все маркетплейсы")
        else:
            logger.info(f"\n✅ Выравнивание не требуется - все артикулы присутствуют")
        
        return dfs
    
    
    def _detect_unit(self, column_name: str) -> Optional[str]:
        """
        Определяет единицу измерения из названия столбца
        
        Args:
            column_name: название столбца
        
        Returns:
            Единица измерения ('kg', 'g', 'mm', 'cm') или None
        """
        if not column_name:
            return None
        
        column_lower = column_name.lower()
        
        # Определяем единицы веса
        if 'кг' in column_lower or 'kg' in column_lower:
            return 'kg'
        if ' г' in column_lower or ',г' in column_lower or 'gram' in column_lower or column_lower.endswith('г'):
            return 'g'
        
        # Определяем единицы длины/размера
        if 'мм' in column_lower or 'mm' in column_lower:
            return 'mm'
        if 'см' in column_lower or 'cm' in column_lower:
            return 'cm'
        
        return None
    
    def _convert_value(
        self, 
        value, 
        from_unit: Optional[str], 
        to_unit: Optional[str]
    ):
        """
        Конвертирует значение между единицами измерения
        
        Args:
            value: исходное значение
            from_unit: исходная единица измерения
            to_unit: целевая единица измерения
        
        Returns:
            Сконвертированное значение
        """
        # Если единицы измерения не определены или одинаковые - возвращаем как есть
        if not from_unit or not to_unit or from_unit == to_unit:
            return value
        
        # Если значение пустое или не числовое - возвращаем как есть
        if pd.isna(value):
            return value
        
        try:
            numeric_value = float(value)
        except (ValueError, TypeError):
            return value
        
        # Конвертация веса
        if from_unit == 'kg' and to_unit == 'g':
            result = numeric_value * 1000
            print(f"      [Конвертация] {numeric_value} кг → {result} г")
            return result
        elif from_unit == 'g' and to_unit == 'kg':
            result = numeric_value / 1000
            print(f"      [Конвертация] {numeric_value} г → {result} кг")
            return result
        
        # Конвертация размеров
        elif from_unit == 'mm' and to_unit == 'cm':
            result = numeric_value / 10
            print(f"      [Конвертация] {numeric_value} мм → {result} см")
            return result
        elif from_unit == 'cm' and to_unit == 'mm':
            result = numeric_value * 10
            print(f"      [Конвертация] {numeric_value} см → {result} мм")
            return result
        
        # Если конвертация не поддерживается - возвращаем исходное значение
        return value
    
    def synchronize_data(
        self,
        file_paths: Dict[str, str],
        output_paths: Dict[str, str] = None,
        report_path: str = None
    ) -> Tuple[Dict[str, pd.DataFrame], Dict]:
        logger.info("="*60)
        logger.info("СИНХРОНИЗАЦИЯ ДАННЫХ МЕЖДУ МАРКЕТПЛЕЙСАМИ")
        logger.info("="*60)

        # 1. Загружаем данные из всех трех файлов
        dfs = self._load_all_dataframes(file_paths)
        
        # 2. Выравниваем артикулы
        dfs = self._align_articles(dfs)

        # 3. 🆕 СИНХРОНИЗАЦИЯ ГАБАРИТОВ (ПЕРЕД остальными столбцами!)
        dimensions_synced = DimensionsSynchronizer.sync_dimensions(dfs)

        # 4. Синхронизируем ОСТАЛЬНЫЕ данные
        logger.info("\n" + "="*60)
        logger.info("📝 СИНХРОНИЗАЦИЯ ОСТАЛЬНЫХ СТОЛБЦОВ")
        logger.info("="*60)
        synced_dfs = self._sync_all_matches(dfs)

        # 5. Сохраняем результаты
        if output_paths:
            self._save_results(synced_dfs, output_paths)

        logger.info("\n" + "="*60)
        logger.info("✅ СИНХРОНИЗАЦИЯ ЗАВЕРШЕНА!")
        logger.info("="*60)
        
        return synced_dfs, self.changes_log
        
    
    def _load_all_dataframes(self, file_paths: Dict[str, str]) -> Dict[str, pd.DataFrame]:
        """Загружает данные через openpyxl для сохранения форматов"""
        logger.info("📂 Загружаю данные из файлов...")
        
        dfs = {}
        # 🆕 Словарь для хранения оригинальных названий столбцов
        self.original_column_names = {}
        
        for marketplace, file_path in file_paths.items():
            self.original_file_paths[marketplace] = file_path
            config = FILE_CONFIGS[marketplace]
            
            wb = load_workbook(file_path, data_only=True)
            ws = wb[config['sheet_name']]
            
            self._load_column_validations(ws, marketplace, config)
            
            data = []
            headers = []
            
            # Читаем заголовки
            for cell in ws[config['header_row']]:
                headers.append(cell.value if cell.value else '')
            
            # 🆕 ОБРАБОТКА ДУБЛИКАТОВ СТОЛБЦОВ
            original_headers = headers.copy()
            seen = {}
            renamed_columns = {}
            
            for i, col in enumerate(headers):
                if col in seen:
                    # Нашли дубликат - добавляем суффикс
                    seen[col] += 1
                    new_name = f"{col}{seen[col]}"
                    logger.warning(f"⚠️ [{marketplace}] Дубликат столбца '{col}' переименован в '{new_name}'")
                    headers[i] = new_name
                    renamed_columns[new_name] = col  # Сохраняем маппинг для возврата
                else:
                    seen[col] = 0
            
            # Сохраняем информацию для восстановления
            if renamed_columns:
                self.original_column_names[marketplace] = {
                    'renamed': renamed_columns,
                    'all_headers': original_headers
                }
            
            # Используем data_start_row вместо header_row + 1
            data_start = config.get('data_start_row', config['header_row'] + 1)
            
            # Читаем данные
            for row in ws.iter_rows(min_row=data_start, values_only=True):
                data.append(row)
            
            df = pd.DataFrame(data, columns=headers)
            dfs[marketplace] = df
            
            wb.close()
            logger.info(f"✅ {config['display_name']}: загружено {len(df)} товаров")
        
        
        return dfs


    def _load_column_validations(self, ws, marketplace: str, config: Dict):
        """
        Загружает информацию о validation для каждого столбца
        """
        from openpyxl.utils import range_boundaries
        
        if marketplace not in self.column_validations:
            self.column_validations[marketplace] = {}
        
        header_row = config['header_row']
        
        # Создаем маппинг: номер колонки -> название
        col_idx_to_name = {}
        for col_idx, cell in enumerate(ws[header_row], start=1):
            if cell.value:
                col_name = str(cell.value).strip()
                col_idx_to_name[col_idx] = col_name
        
        logger.info(f"📋 [{marketplace}] Найдено {len(col_idx_to_name)} столбцов")
        logger.debug(f"[{marketplace}] Первые 5 столбцов: {list(col_idx_to_name.values())[:5]}")
        
        # ДОБАВЬТЕ: Получаем именованные диапазоны
        workbook = ws.parent
        named_ranges = {}
        
        try:
            # ПРАВИЛЬНЫЙ способ итерации по именованным диапазонам
            for name_obj in workbook.defined_names.values():
                try:
                    if name_obj.value:
                        named_ranges[name_obj.name] = name_obj.value
                except Exception as e:
                    logger.debug(f"[{marketplace}] Пропущен именованный диапазон: {e}")
            
            logger.info(f"[{marketplace}] Найдено {len(named_ranges)} именованных диапазонов")
            
            # Выводим первые 5 для проверки
            if named_ranges:
                sample = list(named_ranges.items())[:5]
                for name, value in sample:
                    logger.debug(f"[{marketplace}] Именованный диапазон '{name}' = '{value}'")
        except Exception as e:
            logger.error(f"[{marketplace}] Ошибка получения именованных диапазонов: {e}")
        
        # Проходим по всем validation правилам
        validation_count = 0
        dv_index = 0
        
        for dv in ws.data_validations.dataValidation:
            dv_index += 1
            logger.debug(f"[{marketplace}] DV #{dv_index}: type={dv.type}, sqref={dv.sqref}")
            
            if dv.type != "list":
                logger.debug(f"[{marketplace}] DV #{dv_index}: ПРОПУЩЕН (type != 'list')")
                continue
                
            if dv.sqref is None:
                logger.debug(f"[{marketplace}] DV #{dv_index}: ПРОПУЩЕН (sqref is None)")
                continue
            
            # Извлекаем значения из validation
            allowed_values = []
            if dv.formula1:
                formula = dv.formula1
                logger.debug(f"[{marketplace}] DV #{dv_index}: formula1='{formula[:100]}'...")
                
                # Список задан прямо: "Красный,Синий,Зеленый"
                if formula.startswith('"') and formula.endswith('"'):
                    allowed_values = [v.strip() for v in formula.strip('"').split(',')]
                    logger.debug(f"[{marketplace}] DV #{dv_index}: Прямой список, {len(allowed_values)} значений")
                
                # ДОБАВЬТЕ: Проверяем именованный диапазон
                elif formula in named_ranges:
                    logger.debug(f"[{marketplace}] DV #{dv_index}: Именованный диапазон '{formula}'")
                    try:
                        range_formula = named_ranges[formula]
                        logger.debug(f"[{marketplace}] DV #{dv_index}: Формула диапазона: '{range_formula}'")
                        
                        # Парсим формулу вида "Лист!$A$1:$A$10"
                        clean_formula = range_formula.replace('$', '')
                        if '!' in clean_formula:
                            sheet_name, range_ref = clean_formula.split('!', 1)
                            # Убираем кавычки из имени листа если есть
                            sheet_name = sheet_name.strip("'")
                            target_ws = workbook[sheet_name]
                        else:
                            range_ref = clean_formula
                            target_ws = ws
                        
                        # Извлекаем значения
                        for row in target_ws[range_ref]:
                            for cell in row:
                                if cell.value is not None:
                                    allowed_values.append(str(cell.value).strip())
                        
                        logger.info(f"✅ [{marketplace}] DV #{dv_index}: Извлечено {len(allowed_values)} значений из именованного диапазона '{formula}'")
                    except Exception as e:
                        logger.error(f"[{marketplace}] DV #{dv_index}: Ошибка обработки именованного диапазона '{formula}': {e}")
                
                # Список задан через обычный диапазон
                elif ':' in formula:
                    try:
                        clean_formula = formula.replace('$', '')
                        if '!' in clean_formula:
                            sheet_name, range_ref = clean_formula.split('!')
                            target_ws = workbook[sheet_name]
                            logger.debug(f"[{marketplace}] DV #{dv_index}: Диапазон на листе '{sheet_name}': {range_ref}")
                        else:
                            range_ref = clean_formula
                            target_ws = ws
                            logger.debug(f"[{marketplace}] DV #{dv_index}: Диапазон на текущем листе: {range_ref}")
                        
                        for row in target_ws[range_ref]:
                            for cell in row:
                                if cell.value is not None:
                                    allowed_values.append(str(cell.value).strip())
                        
                        logger.debug(f"[{marketplace}] DV #{dv_index}: Извлечено {len(allowed_values)} значений")
                    except Exception as e:
                        logger.error(f"[{marketplace}] DV #{dv_index}: Ошибка извлечения validation: {e}")
                else:
                    logger.warning(f"[{marketplace}] DV #{dv_index}: Неизвестный формат формулы: '{formula}'")
            else:
                logger.debug(f"[{marketplace}] DV #{dv_index}: formula1 отсутствует")
            
            if not allowed_values:
                logger.debug(f"[{marketplace}] DV #{dv_index}: ПРОПУЩЕН (пустой список значений)")
                continue
            
            # Определяем какие столбцы затронуты этим validation
            ranges = str(dv.sqref).split()
            logger.debug(f"[{marketplace}] DV #{dv_index}: sqref содержит {len(ranges)} диапазон(ов): {ranges}")
            
            for range_str in ranges:
                try:
                    if ':' in range_str:
                        min_col, min_row, max_col, max_row = range_boundaries(range_str)
                        logger.debug(f"[{marketplace}] DV #{dv_index}: Диапазон {range_str} -> столбцы {min_col}-{max_col}, строки {min_row}-{max_row}")
                        
                        # Применяем validation ко всем колонкам в диапазоне
                        for col_idx in range(min_col, max_col + 1):
                            if col_idx in col_idx_to_name:
                                col_name = col_idx_to_name[col_idx]
                                self.column_validations[marketplace][col_name] = allowed_values
                                validation_count += 1
                                logger.info(f"✅ [{marketplace}] Validation для '{col_name}': {len(allowed_values)} значений")
                            else:
                                logger.debug(f"[{marketplace}] DV #{dv_index}: Столбец {col_idx} не найден в заголовках")
                except Exception as e:
                    logger.error(f"[{marketplace}] DV #{dv_index}: Ошибка обработки range_str '{range_str}': {e}")
        
        # Итоговая статистика
        logger.info(f"📊 [{marketplace}] Итого загружено validation для {validation_count} столбцов из {len(ws.data_validations.dataValidation)} правил")
        
        if validation_count == 0:
            logger.warning(f"⚠️ [{marketplace}] НЕ НАЙДЕНО ни одного validation!")
        
        for col_name, values in self.column_validations.get(marketplace, {}).items():
            logger.debug(f"  • {col_name}: {len(values)} значений (первые 3: {values[:3]})")

    
    def _sync_all_matches(self, dfs: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
        """Синхронизирует все совпадающие столбцы"""
        
        # Создаем копии для работы
        synced_dfs = {
            'wildberries': dfs['wildberries'].copy(),
            'ozon': dfs['ozon'].copy(),
            'yandex': dfs['yandex'].copy()
        }
        
        # 🆕 НОВОЕ: Выравниваем артикулы ПЕРЕД синхронизацией
        synced_dfs = self._align_articles(synced_dfs)
        
        # Синхронизируем совпадения всех трех маркетплейсов
        print("\n[*] Синхронизирую совпадения всех 3 маркетплейсов...")
        synced_dfs = self._sync_three_way_matches(synced_dfs)
        
        # Синхронизируем совпадения между двумя маркетплейсами
        print("\n[*] Синхронизирую совпадения между парами маркетплейсов...")
        synced_dfs = self._sync_two_way_matches(synced_dfs)
        
        return synced_dfs
    
    def _sync_three_way_matches(self, dfs: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
        """Синхронизирует совпадения всех трех маркетплейсов"""
        matches = self.comparison_result.get('matches_all_three', [])
        if not matches:
            print("  Нет совпадений для синхронизации")
            return dfs
        
        total_filled = 0
        skipped_count = 0
        
        for match in matches:
            col_wb = match.get('column_1')
            col_ozon = match.get('column_2')
            col_yandex = match.get('column_3')
            
            if not all([col_wb, col_ozon, col_yandex]):
                continue
            
            # Пропускаем исключенные столбцы
            if (is_excluded_column(col_wb) or
                is_excluded_column(col_ozon) or
                is_excluded_column(col_yandex)):
                skipped_count += 1
                continue

            # ⭐ НОВОЕ: Пропускаем габариты - они обрабатываются через DimensionsSynchronizer
            dimensions_columns = {
                'Длина упаковки (целое число)',
                'Ширина упаковки (целое число)', 
                'Высота упаковки (целое число)',
                'Длина упаковки, мм*',
                'Ширина упаковки, мм*',
                'Высота упаковки, мм*',
                'Габариты с упаковкой, см'
            }

            if col_wb in dimensions_columns or col_ozon in dimensions_columns or col_yandex in dimensions_columns:
                skipped_count += 1
                logger.info(f"⏭️ Пропущено (габариты): {col_wb} ↔ {col_ozon} ↔ {col_yandex}")
                continue
            
            # ⭐ ДОБАВЬ ЭТУ ПРОВЕРКУ:
            # Пропускаем габариты - они обрабатываются через DimensionsSynchronizer
            # if col_yandex == "Габариты с упаковкой, см":
            #     skipped_count += 1
            #     logger.info(f"⏭️  Пропущено (габариты): {col_wb} ↔ {col_ozon} ↔ {col_yandex}")
            #     continue
            
            # Проверяем, что столбцы существуют
            if (col_wb not in dfs['wildberries'].columns or
                col_ozon not in dfs['ozon'].columns or
                col_yandex not in dfs['yandex'].columns):
                continue
            
            # Синхронизируем данные между тремя файлами
            filled = self._sync_three_columns(
                dfs,
                col_wb, col_ozon, col_yandex
            )
            
            if filled > 0:
                confidence = int(match.get('confidence', 0) * 100)
                print(f"  ✓ Заполнено {filled} значений: '{col_wb}' ↔ '{col_ozon}' ↔ '{col_yandex}' ({confidence}%)")
                total_filled += filled
        
        if skipped_count > 0:
            print(f"[!] Пропущено {skipped_count} исключенных столбцов")
        print(f"[+] Всего заполнено {total_filled} пустых ячеек в совпадениях всех 3 маркетплейсов")
        
        return dfs
    
    def _sync_two_way_matches(self, dfs: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
        """Синхронизирует совпадения между парами маркетплейсов"""
        
        pairs = [
            ('matches_1_2', 'wildberries', 'ozon', 'column_1', 'column_2'),
            ('matches_1_3', 'wildberries', 'yandex', 'column_1', 'column_3'),
            ('matches_2_3', 'ozon', 'yandex', 'column_2', 'column_3')
        ]
        
        total_filled = 0
        skipped_count = 0
        
        for match_key, mp1, mp2, col_key1, col_key2 in pairs:
            matches = self.comparison_result.get(match_key, [])
            
            if not matches:
                continue
            
            for match in matches:
                col1 = match.get(col_key1)
                col2 = match.get(col_key2)
                
                if not all([col1, col2]):
                    continue
                
                # Пропускаем исключенные столбцы
                if is_excluded_column(col1) or is_excluded_column(col2):
                    skipped_count += 1
                    continue

                # ⭐ НОВОЕ: Пропускаем габариты
                dimensions_columns = {
                    'Длина упаковки (целое число)',
                    'Ширина упаковки (целое число)',
                    'Высота упаковки (целое число)',
                    'Длина упаковки, мм*',
                    'Ширина упаковки, мм*',
                    'Высота упаковки, мм*',
                    'Габариты с упаковкой, см'
                }

                if col1 in dimensions_columns or col2 in dimensions_columns:
                    skipped_count += 1
                    logger.info(f"⏭️ Пропущено (габариты): {mp1}:{col1} ↔ {mp2}:{col2}")
                    continue
                
                # Проверяем, что столбцы существуют
                if col1 not in dfs[mp1].columns or col2 not in dfs[mp2].columns:
                    continue
                
                # Синхронизируем данные между двумя файлами
                filled = self._sync_two_columns(dfs, mp1, mp2, col1, col2)
                
                if filled > 0:
                    confidence = int(match.get('confidence', 0) * 100)
                    print(f"  ✓ Заполнено {filled} значений: {mp1}:'{col1}' ↔ {mp2}:'{col2}' ({confidence}%)")
                    total_filled += filled
        
        if skipped_count > 0:
            print(f"[!] Пропущено {skipped_count} исключенных столбцов")
        print(f"[+] Всего заполнено {total_filled} пустых ячеек в совпадениях между парами")
        return dfs
    
    def _sync_three_columns(
        self,
        dfs: Dict[str, pd.DataFrame],
        col_wb: str,
        col_ozon: str,
        col_yandex: str
    ) -> int:
        """
        Синхронизирует данные между тремя столбцами на основе артикулов
        Returns:
            Количество заполненных ячеек
        """
        filled_count = 0
        
        # Определяем единицы измерения для каждого столбца
        unit_wb = self._detect_unit(col_wb)
        unit_ozon = self._detect_unit(col_ozon)
        unit_yandex = self._detect_unit(col_yandex)
        
        # Создаем словари для быстрого поиска по артикулу
        wb_data = self._create_article_map(dfs['wildberries'], self.article_columns['wildberries'], col_wb)
        ozon_data = self._create_article_map(dfs['ozon'], self.article_columns['ozon'], col_ozon)
        yandex_data = self._create_article_map(dfs['yandex'], self.article_columns['yandex'], col_yandex)
        
        # Получаем все уникальные артикулы
        all_articles = set(wb_data.keys()) | set(ozon_data.keys()) | set(yandex_data.keys())
        
        for article in all_articles:
            if not article:  # Пропускаем пустые артикулы
                continue
            
            # Получаем значения из всех трех источников
            values = {
                'wildberries': wb_data.get(article, {}).get('value'),
                'ozon': ozon_data.get(article, {}).get('value'),
                'yandex': yandex_data.get(article, {}).get('value')
            }
            
            # Находим непустое значение и его источник
            source_value = None
            source_unit = None
            for marketplace, val in values.items():
                # ИСПРАВЛЕНИЕ: проверяем тип и извлекаем скаляр
                if isinstance(val, pd.Series):
                    if not val.empty:
                        val = val.iloc[0]  # Берем первое значение
                    else:
                        val = None
                
                if pd.notna(val) and str(val).strip():
                    source_value = val
                    if marketplace == 'wildberries':
                        source_unit = unit_wb
                    elif marketplace == 'ozon':
                        source_unit = unit_ozon
                    else:
                        source_unit = unit_yandex
                    break
            
            if source_value is None:
                continue
            
            # WB
            if article in wb_data:
                val_wb = values['wildberries']
                if isinstance(val_wb, pd.Series):
                    val_wb = val_wb.iloc[0] if not val_wb.empty else None
                
                if pd.isna(val_wb) or not str(val_wb).strip():
                    idx = wb_data[article]['index']
                    series = dfs['wildberries'][col_wb]
                    if isinstance(series, pd.DataFrame):
                        series = series.iloc[:, 0]  # Берем первый столбец если это DataFrame
                    col_dtype = series.dtype
                    converted_value = self._convert_value(source_value, source_unit, unit_wb)
                    
                    # Проверка validation через AI
                    final_value = self._validate_multiple_values(converted_value, 'wildberries', col_wb)
                    
                    try:
                        # ИСПРАВЛЕНИЕ:
                        if final_value:
                            value_to_set = final_value
                        elif not self.column_validations.get('wildberries', {}).get(col_wb):
                            # Нет validation - записываем как есть
                            value_to_set = converted_value
                        else:
                            # Есть validation но совпадение не найдено - НЕ записываем
                            logger.warning(f"⚠️ [WB] Пропущено '{converted_value}' для '{col_wb}' (не прошло validation)")
                            continue
                        
                        if pd.api.types.is_numeric_dtype(col_dtype):
                            value_to_set = pd.to_numeric(value_to_set, errors='coerce')
                        dfs['wildberries'].at[idx, col_wb] = value_to_set
                        filled_count += 1
                        self._log_change('wildberries', article, col_wb, value_to_set, source_marketplace='ozon' if source_unit == unit_ozon else ('yandex' if source_unit == unit_yandex else 'wildberries'))
                    except Exception:
                        pass
            
            # OZON
            if article in ozon_data:
                val_ozon = values['ozon']
                if isinstance(val_ozon, pd.Series):
                    val_ozon = val_ozon.iloc[0] if not val_ozon.empty else None
                
                if pd.isna(val_ozon) or not str(val_ozon).strip():
                    idx = ozon_data[article]['index']
                    series = dfs['ozon'][col_ozon]
                    if isinstance(series, pd.DataFrame):
                        series = series.iloc[:, 0]
                    col_dtype = series.dtype
                    converted_value = self._convert_value(source_value, source_unit, unit_ozon)
                    
                    final_value = self._validate_multiple_values(converted_value, 'ozon', col_ozon)
                    
                    try:
                        # ИСПРАВЛЕНИЕ:
                        if final_value:
                            value_to_set = final_value
                        elif not self.column_validations.get('ozon', {}).get(col_ozon):
                            value_to_set = converted_value
                        else:
                            logger.warning(f"⚠️ [OZON] Пропущено '{converted_value}' для '{col_ozon}' (не прошло validation)")
                            continue
                        
                        if pd.api.types.is_numeric_dtype(col_dtype):
                            value_to_set = pd.to_numeric(value_to_set, errors='coerce')
                        dfs['ozon'].at[idx, col_ozon] = value_to_set
                        filled_count += 1
                        self._log_change('ozon', article, col_ozon, value_to_set, source_marketplace='wildberries' if source_unit == unit_wb else ('yandex' if source_unit == unit_yandex else 'ozon'))
                    except Exception:
                        pass
            
            # YANDEX
            if article in yandex_data:
                val_yandex = values['yandex']
                if isinstance(val_yandex, pd.Series):
                    val_yandex = val_yandex.iloc[0] if not val_yandex.empty else None
                
                if pd.isna(val_yandex) or not str(val_yandex).strip():
                    idx = yandex_data[article]['index']
                    series = dfs['yandex'][col_yandex]
                    if isinstance(series, pd.DataFrame):
                        series = series.iloc[:, 0]
                    col_dtype = series.dtype
                    
                    # ✅ ПРОВЕРКА: Это композитный столбец габаритов?
                    if col_yandex == "Габариты с упаковкой, см":
                        # Формируем композитное значение из трёх измерений
                        # Определяем источник (WB или Ozon)
                        if source_unit == unit_wb and article in wb_data:
                            # Источник - WB
                            wb_row = dfs['wildberries'][dfs['wildberries'][self.article_columns['wildberries']].astype(str).str.strip() == article].iloc[0]
                            length = wb_row.get('Длина упаковки (целое число)', None)
                            width = wb_row.get('Ширина упаковки (целое число)', None)
                            height = wb_row.get('Высота упаковки (целое число)', None)
                            
                            if all(pd.notna(v) for v in [length, width, height]):
                                composite = DimensionsSynchronizer.format_composite_dimensions(
                                    float(length), float(width), float(height)
                                )
                                dfs['yandex'].at[idx, col_yandex] = composite
                                filled_count += 1
                                self._log_change('yandex', article, col_yandex, composite, source_marketplace='wildberries')
                                continue
                                
                        elif source_unit == unit_ozon and article in ozon_data:
                            # Источник - Ozon (конвертируем мм → см)
                            ozon_row = dfs['ozon'][dfs['ozon'][self.article_columns['ozon']].astype(str).str.strip() == article].iloc[0]
                            length_mm = ozon_row.get('Длина упаковки, мм*', None)
                            width_mm = ozon_row.get('Ширина упаковки, мм*', None)
                            height_mm = ozon_row.get('Высота упаковки, мм*', None)
                            
                            if all(pd.notna(v) for v in [length_mm, width_mm, height_mm]):
                                composite = DimensionsSynchronizer.format_composite_dimensions(
                                    DimensionsSynchronizer.mm_to_cm(float(length_mm)),
                                    DimensionsSynchronizer.mm_to_cm(float(width_mm)),
                                    DimensionsSynchronizer.mm_to_cm(float(height_mm))
                                )
                                dfs['yandex'].at[idx, col_yandex] = composite
                                filled_count += 1
                                self._log_change('yandex', article, col_yandex, composite, source_marketplace='ozon')
                                continue
                    
                    # Обычная логика для остальных столбцов
                    converted_value = self._convert_value(source_value, source_unit, unit_yandex)
                    final_value = self._validate_multiple_values(converted_value, 'yandex', col_yandex)
                    
                    try:
                        if final_value:
                            value_to_set = final_value
                        elif not self.column_validations.get('yandex', {}).get(col_yandex):
                            value_to_set = converted_value
                        else:
                            logger.warning(f"⚠️ [YANDEX] Пропущено '{converted_value}' для '{col_yandex}' (не прошло validation)")
                            continue
                        
                        if pd.api.types.is_numeric_dtype(col_dtype):
                            value_to_set = pd.to_numeric(value_to_set, errors='coerce')
                        
                        dfs['yandex'].at[idx, col_yandex] = value_to_set
                        filled_count += 1
                        self._log_change('yandex', article, col_yandex, value_to_set, 
                                    source_marketplace='wildberries' if source_unit == unit_wb else ('ozon' if source_unit == unit_ozon else 'yandex'))
                    except Exception:
                        pass
        
        return filled_count
    
    def _validate_with_ai(self, value, marketplace: str, column_name: str) -> Optional[str]:
        """
        Проверяет значение через AI если есть validation
        Returns:
            Сопоставленное значение или None если нет validation
        """
        # Проверяем есть ли validation для этого столбца
        allowed_values = self.column_validations.get(marketplace, {}).get(column_name)
        
        if not allowed_values or not self.ai_comparator:
            return None
        
        value_str = str(value).strip()
        
        # === МАППИНГ ЧИСЛИТЕЛЬНЫХ ===
        WORD_TO_NUMBER = {
            'ноль': '0', 'нуль': '0',
            'один': '1', 'одна': '1', 'одно': '1',
            'два': '2', 'две': '2',
            'три': '3',
            'четыре': '4',
            'пять': '5',
            'шесть': '6',
            'семь': '7',
            'восемь': '8',
            'девять': '9',
            'десять': '10',
            'одиннадцать': '11',
            'двенадцать': '12',
        }
        
        # Функция нормализации
        def normalize(text: str) -> str:
            """Нормализует текст: нижний регистр, ё→е"""
            return text.lower().replace('ё', 'е').strip()
        
        # Функция извлечения числа
        def extract_number(text: str) -> Optional[str]:
            """Извлекает первое число из строки типа '1 шт', '2 компрессора'"""
            import re
            numbers = re.findall(r'\d+', text)
            return numbers[0] if numbers else None
        
        # Функция логирования
        def log_match(original: str, matched: str, method: str):
            """Записывает успешное сопоставление в лог"""
            self.ai_validation_log.append({
                'Маркетплейс': marketplace.upper(),
                'Столбец': column_name,
                'Исходное значение': original,
                'Сопоставлено с': matched,
                'Метод': method
            })
        
        # 1. Проверяем точное совпадение
        if value_str in allowed_values:
            logger.info(f"[_validate_with_ai] ТОЧНОЕ совпадение: '{value_str}'")
            log_match(value_str, value_str, 'Точное совпадение')
            return value_str
        
        # 2. Проверяем с нормализацией (регистр + ё/е)
        value_normalized = normalize(value_str)
        for allowed in allowed_values:
            if normalize(allowed) == value_normalized:
                logger.info(f"[_validate_with_ai] Совпадение с нормализацией: '{value_str}' → '{allowed}'")
                log_match(value_str, allowed, 'Нормализация (регистр/ё-е)')
                return allowed
        
        # 3. НОВОЕ: Проверяем числительные ("два" → "2")
        if value_normalized in WORD_TO_NUMBER:
            number_value = WORD_TO_NUMBER[value_normalized]
            if number_value in allowed_values:
                logger.info(f"[_validate_with_ai] Числительное: '{value_str}' → '{number_value}'")
                log_match(value_str, number_value, 'Маппинг числительных')
                return number_value
            # Проверяем также с нормализацией allowed_values
            for allowed in allowed_values:
                if normalize(allowed) == number_value or extract_number(allowed) == number_value:
                    logger.info(f"[_validate_with_ai] Числительное (fuzzy): '{value_str}' → '{allowed}'")
                    log_match(value_str, allowed, 'Маппинг числительных')
                    return allowed
        
        # 4. Извлекаем число если это числовое поле
        number = extract_number(value_str)
        if number:
            # Проверяем точное совпадение числа
            if number in allowed_values:
                logger.info(f"[_validate_with_ai] Извлечено число: '{value_str}' → '{number}'")
                log_match(value_str, number, 'Извлечение числа')
                return number
            
            # Проверяем с нормализацией
            for allowed in allowed_values:
                if extract_number(allowed) == number:
                    logger.info(f"[_validate_with_ai] Совпадение по числу: '{value_str}' → '{allowed}'")
                    log_match(value_str, allowed, 'Извлечение числа')
                    return allowed
        
        # 5. Проверяем частичное совпадение (по словам)
        value_words = set(value_normalized.split())
        for allowed in allowed_values:
            allowed_words = set(normalize(allowed).split())
            
            # Если все слова из value есть в allowed
            if value_words and value_words.issubset(allowed_words):
                logger.info(f"[_validate_with_ai] Частичное совпадение: '{value_str}' → '{allowed}'")
                log_match(value_str, allowed, 'Частичное совпадение (слова)')
                return allowed
        
        # 6. НОВОЕ: Проверяем кэш перед AI запросом
        cache_key = (value_str, column_name)
        if cache_key in self.ai_cache:
            cached_result = self.ai_cache[cache_key]
            if cached_result:
                logger.info(f"📦 [CACHE] Найдено в кэше: '{value_str}' → '{cached_result}'")
                log_match(value_str, cached_result, 'Кэш AI')
            else:
                logger.info(f"📦 [CACHE] Найдено в кэше: '{value_str}' → НЕТ СОВПАДЕНИЯ")
            return cached_result
        
        # 7. Спрашиваем AI (и сохраняем в кэш)
        logger.info(f"🤖 [AI] Проверяю '{value_str}' для столбца '{column_name}'...")
        matched_value = self.ai_comparator.match_value_with_list(value_str, allowed_values, column_name=column_name)
        
        # Сохраняем в кэш (даже если None)
        self.ai_cache[cache_key] = matched_value
        
        if matched_value:
            logger.info(f"✅ [AI] Найдено: '{value_str}' → '{matched_value}'")
            log_match(value_str, matched_value, 'AI запрос')
            return matched_value
        else:
            logger.warning(f"❌ [AI] Не найдено совпадение для '{value_str}'")
            return None
    
    def _validate_multiple_values(self, value, marketplace: str, column_name: str) -> Optional[str]:
        """
        Валидирует значения с разделителями (;) и форматирует согласно требованиям маркетплейса
        
        Args:
            value: исходное значение (может содержать ";")
            marketplace: 'wildberries', 'ozon', 'yandex'
            column_name: название столбца
        
        Returns:
            Отформатированная строка или None
        """
        if not value:
            return None
        
        value_str = str(value).strip()
        
        # Проверяем есть ли разделители
        if ';' not in value_str:
            # Одно значение - обычная валидация
            return self._validate_with_ai(value_str, marketplace, column_name)
        
        # Множественные значения - разбиваем по ";"
        parts = [part.strip() for part in value_str.split(';') if part.strip()]
        
        if not parts:
            return None
        
        # Wildberries: только ПЕРВОЕ значение
        if marketplace == 'wildberries':
            validated = self._validate_with_ai(parts[0], marketplace, column_name)
            return validated if validated else parts[0]  # Если валидация не прошла, берём как есть
        
        # Ozon и Яндекс: валидируем каждое значение
        validated_parts = []
        for part in parts:
            validated = self._validate_with_ai(part, marketplace, column_name)
            if validated and validated not in validated_parts:  # Избегаем дубликатов
                validated_parts.append(validated)
        
        if not validated_parts:
            return None
        
        # Форматируем согласно требованиям маркетплейса
        if marketplace == 'yandex':
            return ', '.join(validated_parts)  # "Красный, Синий"
        elif marketplace == 'ozon':
            return '; '.join(validated_parts)  # "Красный; Синий"
        
        return validated_parts[0]  # На всякий случай


    
    def _sync_two_columns(
        self,
        dfs: Dict[str, pd.DataFrame],
        mp1: str,
        mp2: str,
        col1: str,
        col2: str
    ) -> int:
        """
        Синхронизирует данные между двумя столбцами на основе артикулов
        Returns:
            Количество заполненных ячеек
        """
        filled_count = 0
        
        # Определяем единицы измерения
        unit1 = self._detect_unit(col1)
        unit2 = self._detect_unit(col2)
        
        # Определяем столбцы артикулов
        article_col1 = self.article_columns[mp1]
        article_col2 = self.article_columns[mp2]
        
        # Создаем словари для быстрого поиска
        data1 = self._create_article_map(dfs[mp1], article_col1, col1)
        data2 = self._create_article_map(dfs[mp2], article_col2, col2)
        
        # Получаем все уникальные артикулы
        all_articles = set(data1.keys()) | set(data2.keys())
        
        for article in all_articles:
            if not article:
                continue
            
            # Получаем значения
            val1 = data1.get(article, {}).get('value')
            val2 = data2.get(article, {}).get('value')
            
            # ИСПРАВЛЕНИЕ: проверяем Series
            if isinstance(val1, pd.Series):
                val1 = val1.iloc[0] if not val1.empty else None
            if isinstance(val2, pd.Series):
                val2 = val2.iloc[0] if not val2.empty else None
            
            # Заполняем пустые ячейки
            if article in data1 and article in data2:
                # Если в первом пусто, а во втором есть
                if (pd.isna(val1) or not str(val1).strip()) and pd.notna(val2) and str(val2).strip():
                    idx = data1[article]['index']
                    series = dfs[mp1][col1]
                    if isinstance(series, pd.DataFrame):
                        series = series.iloc[:, 0]
                    col_dtype = series.dtype  # ✅
                    converted_value = self._convert_value(val2, unit2, unit1)
                    
                    final_value = self._validate_multiple_values(converted_value, mp1, col1)
                    
                    try:
                        # ИСПРАВЛЕНИЕ:
                        if final_value:
                            value_to_set = final_value
                        elif not self.column_validations.get(mp1, {}).get(col1):
                            value_to_set = converted_value
                        else:
                            logger.warning(f"⚠️ [{mp1.upper()}] Пропущено '{converted_value}' для '{col1}' (не прошло validation)")
                            continue
                        
                        if pd.api.types.is_numeric_dtype(col_dtype):
                            value_to_set = pd.to_numeric(value_to_set, errors='coerce')
                        dfs[mp1].at[idx, col1] = value_to_set
                        filled_count += 1
                        self._log_change(mp1, article, col1, value_to_set, source_marketplace=mp2)
                    except Exception:
                        pass
                
                # Если во втором пусто, а в первом есть
                elif (pd.isna(val2) or not str(val2).strip()) and pd.notna(val1) and str(val1).strip():
                    idx = data2[article]['index']
                    series = dfs[mp2][col2]
                    if isinstance(series, pd.DataFrame):
                        series = series.iloc[:, 0]
                    col_dtype = series.dtype  # ✅
                    converted_value = self._convert_value(val1, unit1, unit2)
                    
                    final_value = self._validate_multiple_values(converted_value, mp2, col2)
                    
                    try:
                        # ИСПРАВЛЕНИЕ:
                        if final_value:
                            value_to_set = final_value
                        elif not self.column_validations.get(mp2, {}).get(col2):
                            value_to_set = converted_value
                        else:
                            logger.warning(f"⚠️ [{mp2.upper()}] Пропущено '{converted_value}' для '{col2}' (не прошло validation)")
                            continue
                        
                        if pd.api.types.is_numeric_dtype(col_dtype):
                            value_to_set = pd.to_numeric(value_to_set, errors='coerce')
                        dfs[mp2].at[idx, col2] = value_to_set
                        filled_count += 1
                        self._log_change(mp2, article, col2, value_to_set, source_marketplace=mp1)
                    except Exception:
                        pass
        
        return filled_count
    
    def _create_article_map(self, df: pd.DataFrame, article_col: str, value_col: str) -> Dict:
        """
        Создает словарь для быстрого поиска значений по артикулу
        
        Returns:
            Словарь {артикул: {'value': значение, 'index': индекс строки}}
        """
        article_map = {}
        
        if article_col not in df.columns or value_col not in df.columns:
            return article_map
        
        for idx, row in df.iterrows():
            article = row.get(article_col)
            value = row.get(value_col)
            
            if pd.notna(article):
                article_str = str(article).strip()
                if article_str:
                    article_map[article_str] = {
                        'value': value,
                        'index': idx
                    }
        
        return article_map
    
    def _log_change(self, marketplace: str, article: str, column: str, new_value, source_marketplace: str = None):
        """Логирует произведенное изменение"""
        self.changes_log[marketplace].append({
            'article': article,
            'column': column,
            'new_value': str(new_value),
            'source_marketplace': source_marketplace  # ← ДОБАВЛЕНО
        })
    
    def _postprocess_wb_dimensions(self, dfs: Dict[str, pd.DataFrame]) -> int:
        """
        Постобработка габаритов WB: конвертация из мм в см если данные пришли из Ozon
        
        Returns:
            Количество сконвертированных значений
        """
        if 'wildberries' not in dfs:
            return 0
        
        converted_count = 0
        df_wb = dfs['wildberries']
        
        # Столбцы габаритов WB
        wb_dimension_columns = [
            'Длина упаковки (целое число)',
            'Ширина упаковки (целое число)',
            'Высота упаковки (целое число)'
        ]
        
        # Проверяем наличие столбцов
        for col_name in wb_dimension_columns:
            if col_name not in df_wb.columns:
                return 0
        
        # Проходим по всем изменениям WB из Ozon
        if 'wildberries' not in self.changes_log:
            return 0
        
        for change in self.changes_log['wildberries']:
            # Проверяем: это габарит И источник - Ozon
            if (change.get('source_marketplace') == 'ozon' and 
                change.get('column') in wb_dimension_columns):
                
                article = change.get('article')
                column = change.get('column')
                
                # Находим строку с этим артикулом
                mask = df_wb[self.article_columns['wildberries']].astype(str).str.strip() == str(article).strip()
                if mask.any():
                    idx = df_wb[mask].index[0]
                    value = df_wb.at[idx, column]
                    
                    # Конвертируем если >= 100 (защита от повторной конвертации)
                    if pd.notna(value):
                        try:
                            numeric_value = float(value)
                            if numeric_value >= 100:
                                converted_value = round(numeric_value / 10, 1)
                                df_wb.at[idx, column] = converted_value
                                converted_count += 1
                                logger.info(f"  ✓ [{article}] {column}: {numeric_value} мм → {converted_value} см")
                        except (ValueError, TypeError):
                            pass
        
        return converted_count

    
    def _get_validation_list_values(self, ws, row_idx: int, col_idx: int) -> List[str]:
        """
        Получает список допустимых значений из data validation ячейки
        
        Returns:
            Список допустимых значений или пустой список
        """
        from openpyxl.worksheet.datavalidation import DataValidation
        
        cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
        
        # Проходим по всем data validation правилам
        for dv in ws.data_validations.dataValidation:
            if dv.type != "list":
                continue
            
            # ИСПРАВЛЕНИЕ: проверяем принадлежность ячейки к диапазону validation
            # Используем правильную проверку через sqref (string reference)
            if dv.sqref is None:
                continue
                
            # sqref может содержать несколько диапазонов, разделенных пробелами
            # Например: "B2:B100 D2:D100"
            ranges = str(dv.sqref).split()
            
            cell_in_range = False
            for range_str in ranges:
                # Проверяем входит ли наша ячейка в диапазон
                if ':' in range_str:
                    # Диапазон типа A2:A100
                    try:
                        from openpyxl.utils import range_boundaries
                        min_col, min_row, max_col, max_row = range_boundaries(range_str)
                        
                        if (min_col <= col_idx <= max_col and 
                            min_row <= row_idx <= max_row):
                            cell_in_range = True
                            break
                    except:
                        pass
                else:
                    # Одиночная ячейка типа A2
                    if range_str == cell_ref:
                        cell_in_range = True
                        break
            
            if not cell_in_range:
                continue
            
            # Нашли validation для этой ячейки, извлекаем значения
            if dv.formula1:
                formula = dv.formula1
                
                # Список задан прямо: "Красный,Синий,Зеленый"
                if formula.startswith('"') and formula.endswith('"'):
                    values = formula.strip('"').split(',')
                    return [v.strip() for v in values]
                
                # Список задан через диапазон: $A$1:$A$10 или Sheet1!$A$1:$A$10
                elif ':' in formula:
                    try:
                        # Убираем $ для парсинга
                        clean_formula = formula.replace('$', '')
                        
                        # Проверяем, есть ли ссылка на другой лист
                        if '!' in clean_formula:
                            sheet_name, range_ref = clean_formula.split('!')
                            target_ws = ws.parent[sheet_name]
                        else:
                            range_ref = clean_formula
                            target_ws = ws
                        
                        # Извлекаем значения из диапазона
                        values = []
                        for row in target_ws[range_ref]:
                            for cell in row:
                                if cell.value is not None:
                                    values.append(str(cell.value).strip())
                        
                        return values
                        
                    except Exception as e:
                        print(f"      [!] Не удалось извлечь список из диапазона {formula}: {e}")
                        return []
        
        return []
    
    def _create_ai_log_sheet(self, output_paths: Dict[str, str]):
        """Создаёт лист с логами AI-сопоставлений"""
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Создаём DataFrame из логов
        df_log = pd.DataFrame(self.ai_validation_log)
        
        # Добавляем лист в каждый выходной файл
        for marketplace, output_path in output_paths.items():
            try:
                # Загружаем существующий файл
                wb = load_workbook(output_path)
                
                # Создаём новый лист
                if 'AI_Логи' in wb.sheetnames:
                    del wb['AI_Логи']
                ws = wb.create_sheet('AI_Логи', 0)  # 0 = первый лист
                
                # Заголовки
                headers = ['Маркетплейс', 'Столбец', 'Исходное значение', 'Сопоставлено с', 'Метод']
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True, size=11)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Данные
                for row_idx, row_data in enumerate(self.ai_validation_log, start=2):
                    ws.cell(row=row_idx, column=1, value=row_data['Маркетплейс'])
                    ws.cell(row=row_idx, column=2, value=row_data['Столбец'])
                    ws.cell(row=row_idx, column=3, value=row_data['Исходное значение'])
                    ws.cell(row=row_idx, column=4, value=row_data['Сопоставлено с'])
                    ws.cell(row=row_idx, column=5, value=row_data['Метод'])
                
                # Автоширина колонок
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column].width = adjusted_width
                
                # Сохраняем
                wb.save(output_path)
                logger.info(f"✅ Лист 'AI_Логи' добавлен в {output_path}")
                
            except Exception as e:
                logger.error(f"❌ Ошибка при создании AI-лога для {marketplace}: {e}")
    
    def _create_ai_log_sheet_in_report(self, report_path: str):
        """Создаёт лист с логами AI-сопоставлений в файле результата"""
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        try:
            # Загружаем файл результата
            wb = load_workbook(report_path)
            
            # Удаляем старый лист если существует
            if 'AI_Логи' in wb.sheetnames:
                del wb['AI_Логи']
            
            # Создаём новый лист (первым после главного)
            ws = wb.create_sheet('AI_Логи', 1)  # Индекс 1 = второй лист
            
            # Заголовки
            headers = ['Маркетплейс', 'Столбец', 'Исходное значение', 'Сопоставлено с', 'Метод']
            
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Данные
            for row_idx, row_data in enumerate(self.ai_validation_log, start=2):
                ws.cell(row=row_idx, column=1, value=row_data.get('Маркетплейс', ''))
                ws.cell(row=row_idx, column=2, value=row_data.get('Столбец', ''))
                ws.cell(row=row_idx, column=3, value=row_data.get('Исходное значение', ''))
                ws.cell(row=row_idx, column=4, value=row_data.get('Сопоставлено с', ''))
                ws.cell(row=row_idx, column=5, value=row_data.get('Метод', ''))
            
            # Автоширина колонок
            for col in ws.columns:
                max_length = 0
                column_letter = col[0].column_letter
                
                for cell in col:
                    try:
                        cell_length = len(str(cell.value)) if cell.value else 0
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
                
                adjusted_width = min(max_length + 3, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Морозим шапку
            ws.freeze_panes = 'A2'
            
            # Сохраняем
            wb.save(report_path)
            logger.info(f"✅ Лист 'AI_Логи' добавлен в {report_path}")
            logger.info(f"📊 Всего записей AI-логов: {len(self.ai_validation_log)}")
            
        except Exception as e:
            logger.error(f"❌ Ошибка при создании AI-лога: {e}")

    
    def _save_results(self, dfs: Dict[str, pd.DataFrame], output_paths: Dict[str, str]):
        """Сохраняет синхронизированные данные в файлы"""
        print("\n[*] Сохраняю синхронизированные данные...")
        
        stats = {'saved': 0, 'ai_matched': 0, 'validation_conflicts': 0, 'skipped': 0}
        
        for marketplace, df in dfs.items():
            output_path = output_paths.get(marketplace)
            if not output_path:
                continue
            
            config = FILE_CONFIGS[marketplace]
            original_file = self.original_file_paths[marketplace]
            
            print(f"\n[*] Обработка {config['display_name']}...")
            
            # Сбрасываем индексы ПЕРЕД сохранением
            df = df.reset_index(drop=True)
            
            # Открываем ОРИГИНАЛЬНЫЙ файл через openpyxl
            wb = load_workbook(original_file)
            ws = wb[config['sheet_name']]
            
            header_row = config['header_row']
            data_start_row = config.get('data_start_row', header_row + 1)
            
            # 🆕 СОЗДАЁМ МАППИНГ: DataFrame столбец → Excel столбцы (с учётом дубликатов!)
            df_to_excel_mapping = {}  # {df_col_name: [excel_col_1, excel_col_2, ...]}
            header_count = {}
            
            # Читаем заголовки из Excel
            for col_idx, cell in enumerate(ws[header_row], start=1):
                if cell.value:
                    col_name = str(cell.value).strip()
                    
                    # Первое вхождение - оригинальное имя
                    if col_name not in header_count:
                        header_count[col_name] = 0
                        df_col = col_name  # "Вес с упаковкой (кг)"
                    else:
                        # Дубликат - с суффиксом
                        header_count[col_name] += 1
                        df_col = f"{col_name}{header_count[col_name]}"  # "Вес с упаковкой (кг)1"
                    
                    # Сохраняем маппинг
                    if df_col not in df_to_excel_mapping:
                        df_to_excel_mapping[df_col] = []
                    df_to_excel_mapping[df_col].append(col_idx)
            
            # 🆕 СИНХРОНИЗАЦИЯ ДУБЛИКАТОВ ПЕРЕД ЗАПИСЬЮ
            if marketplace in self.original_column_names:
                renamed_map = self.original_column_names[marketplace]['renamed']
                
                for duplicated_name, original_name in renamed_map.items():
                    # duplicated_name = "Вес с упаковкой (кг)1"
                    # original_name = "Вес с упаковкой (кг)"
                    
                    if original_name in df.columns and duplicated_name in df.columns:
                        # Копируем значения из оригинала в дубликат
                        for idx in df.index:
                            original_value = df.at[idx, original_name]
                            if pd.notna(original_value) and str(original_value).strip():
                                df.at[idx, duplicated_name] = original_value
                        
                        logger.info(f"✅ [{marketplace}] Синхронизированы: '{original_name}' → '{duplicated_name}'")
            
            # Расширяем лист если нужно
            current_rows = ws.max_row
            required_rows = data_start_row + len(df)
            
            if required_rows > current_rows:
                print(f"[INFO] Расширяю лист: {current_rows} → {required_rows}")
                last_data_row = current_rows
                for row_idx in range(current_rows + 1, required_rows + 1):
                    for col_idx in range(1, ws.max_column + 1):
                        source_cell = ws.cell(row=last_data_row, column=col_idx)
                        target_cell = ws.cell(row=row_idx, column=col_idx)
                        
                        if source_cell.has_style:
                            target_cell.font = source_cell.font.copy()
                            target_cell.border = source_cell.border.copy()
                            target_cell.fill = source_cell.fill.copy()
                            target_cell.number_format = source_cell.number_format
                            target_cell.protection = source_cell.protection.copy()
                            target_cell.alignment = source_cell.alignment.copy()
            
            # Записываем данные
            for row_num, (df_row_idx, row) in enumerate(df.iterrows()):
                excel_row_idx = data_start_row + row_num
                
                for df_col_name, value in row.items():
                    if pd.isna(value):
                        continue
                    
                    # Получаем список Excel столбцов для этого DataFrame столбца
                    excel_cols = df_to_excel_mapping.get(df_col_name, [])
                    
                    # Записываем во ВСЕ соответствующие столбцы (для дубликатов)
                    for excel_col_idx in excel_cols:
                        cell = ws.cell(row=excel_row_idx, column=excel_col_idx)
                        
                        # Валидация через AI (если есть)
                        allowed_values = self._get_validation_list_values(ws, excel_row_idx, excel_col_idx)
                        
                        if allowed_values and self.ai_comparator:
                            matched_value = self.ai_comparator.match_value_with_list(str(value), allowed_values)
                            if matched_value:
                                cell.value = matched_value
                                stats['ai_matched'] += 1
                            else:
                                stats['validation_conflicts'] += 1
                                stats['skipped'] += 1
                                continue
                        else:
                            cell.value = value
                        
                        stats['saved'] += 1
            
            # Сохраняем файл
            wb.save(output_path)
            print(f"[+] {config['display_name']}: сохранено в '{output_path}'")
        
        # Статистика
        print(f"\n{'='*60}")
        print(f"СТАТИСТИКА:")
        print(f"  ✓ Записано: {stats['saved']}")
        if self.ai_comparator:
            print(f"  🤖 AI-сопоставлений: {stats['ai_matched']}")
            print(f"  ⚠ Конфликтов: {stats['validation_conflicts']}")
        print(f"{'='*60}")
