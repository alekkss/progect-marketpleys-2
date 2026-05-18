"""
Модуль работы с Excel-файлами маркетплейсов.

Отвечает за три связанные задачи:
    1. Загрузка DataFrame из xlsx-файлов с сохранением форматов и validation.
    2. Сохранение синхронизированных данных обратно в файлы.
    3. Создание листа «AI_Логи» с результатами валидации.

Использует openpyxl напрямую (не pandas.read_excel) для сохранения
стилей, формул и data validation из оригинальных файлов МП.

Паттерн: Repository — абстрагирует доступ к xlsx-файлам,
скрывая детали openpyxl от остальных компонентов системы.
Паттерн: Dependency Injection — ai_comparator передаётся через конструктор.
"""

from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config.config import FILE_CONFIGS
from utils.logger_config import setup_logger

logger = setup_logger("excel_io")


class ExcelFileManager:
    """
    Загружает, сохраняет и дополняет Excel-файлы маркетплейсов.

    Принимает зависимости через конструктор (Dependency Inversion):
        - ai_comparator: экземпляр AIComparator или None.

    После загрузки предоставляет:
        - column_validations:   {маркетплейс: {столбец: [допустимые значения]}}.
        - original_column_names: маппинг переименованных дубликатов столбцов.
        - original_file_paths:   пути к исходным файлам для записи.
    """

    def __init__(self, ai_comparator: object = None) -> None:
        """
        Args:
            ai_comparator: экземпляр AIComparator для валидации при сохранении.
                           Может быть None — тогда запись идёт без AI-проверки.
        """
        self._ai_comparator = ai_comparator

        # Заполняются при load_all_dataframes
        self.column_validations: Dict[str, Dict[str, List[str]]] = {}
        self.original_column_names: Dict[str, Dict] = {}
        self.original_file_paths: Dict[str, str] = {}

    # ------------------------------------------------------------------
    # Публичный интерфейс
    # ------------------------------------------------------------------

    def load_all_dataframes(
        self, file_paths: Dict[str, str]
    ) -> Dict[str, pd.DataFrame]:
        """
        Загружает DataFrame из xlsx-файлов через openpyxl.

        Использует openpyxl вместо pandas.read_excel, чтобы сохранить
        data validation, стили и точные значения ячеек.
        Обрабатывает дублирующиеся заголовки: добавляет числовой суффикс
        («Вес (кг)», «Вес (кг)1») и сохраняет маппинг для обратного
        восстановления при сохранении.

        Args:
            file_paths: словарь {маркетплейс: путь к файлу}.

        Returns:
            Словарь {маркетплейс: DataFrame}.
        """
        logger.info("📂 Загружаю данные из файлов...")

        dfs: Dict[str, pd.DataFrame] = {}
        self.original_column_names = {}
        self.column_validations = {}

        for marketplace, file_path in file_paths.items():
            self.original_file_paths[marketplace] = file_path
            config = FILE_CONFIGS[marketplace]

            wb = load_workbook(file_path, data_only=True)
            ws = wb[config["sheet_name"]]

            self._load_column_validations(ws, marketplace, config)

            headers = self._read_headers(ws, config["header_row"])
            headers, renamed_columns = self._deduplicate_headers(
                headers, marketplace
            )

            if renamed_columns:
                original_headers = [
                    renamed_columns.get(h, h) for h in headers
                ]
                self.original_column_names[marketplace] = {
                    "renamed":     renamed_columns,
                    "all_headers": original_headers,
                }

            data_start = config.get(
                "data_start_row", config["header_row"] + 1
            )
            data = [
                row for row in ws.iter_rows(
                    min_row=data_start, values_only=True
                )
            ]

            dfs[marketplace] = pd.DataFrame(data, columns=headers)
            wb.close()

            logger.info(
                f"✅ {config['display_name']}: "
                f"загружено {len(dfs[marketplace])} товаров"
            )

        return dfs

    async def save_results(
        self,
        dfs: Dict[str, pd.DataFrame],
        output_paths: Dict[str, str],
        ai_validation_log: List[Dict],
    ) -> None:
        """
        Сохраняет синхронизированные DataFrame обратно в xlsx-файлы.

        Открывает оригинальный файл, расширяет лист при необходимости,
        записывает данные с учётом дубликатов столбцов и сохраняет.

        Args:
            dfs:               словарь {маркетплейс: DataFrame}.
            output_paths:      словарь {маркетплейс: путь к выходному файлу}.
            ai_validation_log: список записей AI-валидации (принимается для
                               совместимости сигнатуры, лог в МП-файлы не пишется).
        """
        logger.info("\n[*] Сохраняю синхронизированные данные...")

        stats = {
            "saved": 0, "ai_matched": 0,
            "validation_conflicts": 0, "skipped": 0,
        }

        for marketplace, df in dfs.items():
            output_path = output_paths.get(marketplace)
            if not output_path:
                continue

            config = FILE_CONFIGS[marketplace]
            original_file = self.original_file_paths[marketplace]

            logger.info(f"\n[*] Обработка {config['display_name']}...")

            df = df.reset_index(drop=True)

            wb = load_workbook(original_file)
            ws = wb[config["sheet_name"]]

            header_row     = config["header_row"]
            data_start_row = config.get("data_start_row", header_row + 1)

            df_to_excel_mapping = self._build_excel_column_mapping(
                ws, header_row
            )
            self._sync_duplicate_columns(df, marketplace)
            self._expand_sheet_if_needed(ws, data_start_row, len(df))

            await self._write_dataframe_to_sheet(
                ws, df, data_start_row,
                df_to_excel_mapping, marketplace, stats
            )

            wb.save(output_path)
            logger.info(
                f"[+] {config['display_name']}: сохранено в '{output_path}'"
            )

        self._log_save_stats(stats)

    def create_ai_log_in_report(
        self, report_path: str, ai_validation_log: List[Dict]
    ) -> None:
        """
        Создаёт лист «AI_Логи» в отдельном файле отчёта.

        Args:
            report_path:       путь к файлу отчёта.
            ai_validation_log: список записей AI-валидации.
        """
        self._write_ai_log_sheet(
            report_path, ai_validation_log, sheet_index=1
        )
        logger.info(
            f"📊 AI_Логи в отчёте: {len(ai_validation_log)} записей"
        )

    # ------------------------------------------------------------------
    # Загрузка заголовков и validation
    # ------------------------------------------------------------------

    @staticmethod
    def _read_headers(ws, header_row: int) -> List[str]:
        """
        Читает заголовки из строки Excel-листа.

        Args:
            ws:         лист openpyxl.
            header_row: номер строки с заголовками.

        Returns:
            Список заголовков (пустые ячейки заменяются пустой строкой).
        """
        return [
            cell.value if cell.value else ""
            for cell in ws[header_row]
        ]

    @staticmethod
    def _deduplicate_headers(
        headers: List[str], marketplace: str
    ) -> Tuple[List[str], Dict[str, str]]:
        """
        Устраняет дубликаты заголовков, добавляя числовой суффикс.

        Возвращает обновлённый список заголовков и маппинг
        {новое_имя: оригинальное_имя} для последующего восстановления.

        Args:
            headers:     список заголовков.
            marketplace: название МП (для логирования).

        Returns:
            Кортеж (обновлённые заголовки, маппинг переименований).
        """
        seen: Dict[str, int] = {}
        renamed: Dict[str, str] = {}
        result = list(headers)

        for i, col in enumerate(result):
            if col in seen:
                seen[col] += 1
                new_name = f"{col}{seen[col]}"
                logger.warning(
                    f"⚠️ [{marketplace}] Дубликат '{col}' → '{new_name}'"
                )
                result[i] = new_name
                renamed[new_name] = col
            else:
                seen[col] = 0

        return result, renamed

    def _load_column_validations(
        self, ws, marketplace: str, config: Dict
    ) -> None:
        """
        Загружает validation-списки для всех столбцов листа.

        Поддерживает три формата formula1 data validation:
            1. Прямой список: «"Красный,Синий,Зеленый"».
            2. Именованный диапазон: «MyRange».
            3. Диапазон ячеек: «Sheet1!$A$1:$A$10» или «$A$1:$A$10».

        Args:
            ws:          лист openpyxl.
            marketplace: ключ маркетплейса.
            config:      конфигурация МП из FILE_CONFIGS.
        """
        from openpyxl.utils import range_boundaries

        if marketplace not in self.column_validations:
            self.column_validations[marketplace] = {}

        header_row = config["header_row"]
        col_idx_to_name = {
            col_idx: str(cell.value).strip()
            for col_idx, cell in enumerate(ws[header_row], start=1)
            if cell.value
        }

        logger.info(
            f"📋 [{marketplace}] Найдено {len(col_idx_to_name)} столбцов"
        )

        # Собираем именованные диапазоны рабочей книги
        named_ranges = self._collect_named_ranges(ws.parent, marketplace)

        validation_count = 0

        for dv_index, dv in enumerate(
            ws.data_validations.dataValidation, start=1
        ):
            if dv.type != "list" or dv.sqref is None:
                continue

            allowed_values = self._extract_allowed_values(
                dv, ws, named_ranges, marketplace, dv_index
            )
            if not allowed_values:
                continue

            # Применяем validation ко всем столбцам в диапазоне
            for range_str in str(dv.sqref).split():
                try:
                    if ":" in range_str:
                        min_col, _, max_col, _ = range_boundaries(range_str)
                        for col_idx in range(min_col, max_col + 1):
                            if col_idx in col_idx_to_name:
                                col_name = col_idx_to_name[col_idx]
                                self.column_validations[marketplace][col_name] = (
                                    allowed_values
                                )
                                validation_count += 1
                                logger.info(
                                    f"✅ [{marketplace}] Validation '{col_name}': "
                                    f"{len(allowed_values)} значений"
                                )
                except Exception as e:
                    logger.error(
                        f"[{marketplace}] DV #{dv_index}: "
                        f"ошибка диапазона '{range_str}': {e}"
                    )

        logger.info(
            f"📊 [{marketplace}] Загружено validation "
            f"для {validation_count} столбцов"
        )
        if validation_count == 0:
            logger.warning(f"⚠️ [{marketplace}] Не найдено ни одного validation!")

    @staticmethod
    def _collect_named_ranges(
        workbook, marketplace: str
    ) -> Dict[str, str]:
        """
        Собирает именованные диапазоны из рабочей книги.

        Args:
            workbook:    объект Workbook openpyxl.
            marketplace: название МП (для логирования).

        Returns:
            Словарь {имя_диапазона: формула}.
        """
        named_ranges: Dict[str, str] = {}
        try:
            for name_obj in workbook.defined_names.values():
                try:
                    if name_obj.value:
                        named_ranges[name_obj.name] = name_obj.value
                except Exception as e:
                    logger.debug(
                        f"[{marketplace}] Пропущен именованный диапазон: {e}"
                    )
            logger.info(
                f"[{marketplace}] Найдено {len(named_ranges)} "
                f"именованных диапазонов"
            )
        except Exception as e:
            logger.error(
                f"[{marketplace}] Ошибка получения именованных диапазонов: {e}"
            )
        return named_ranges

    @staticmethod
    def _extract_allowed_values(
        dv,
        ws,
        named_ranges: Dict[str, str],
        marketplace: str,
        dv_index: int,
    ) -> List[str]:
        """
        Извлекает допустимые значения из правила data validation.

        Поддерживаемые форматы formula1:
            - «"Красный,Синий"» — прямой список через запятую.
            - «MyRange»         — именованный диапазон.
            - «Sheet1!$A$1:$A$10» или «$A$1:$A$10» — диапазон ячеек.

        Args:
            dv:           объект DataValidation openpyxl.
            ws:           лист openpyxl.
            named_ranges: словарь именованных диапазонов рабочей книги.
            marketplace:  название МП (для логирования).
            dv_index:     порядковый номер правила (для логирования).

        Returns:
            Список допустимых значений или пустой список.
        """
        if not dv.formula1:
            return []

        formula = dv.formula1

        # Прямой список: "Красный,Синий,Зеленый"
        if formula.startswith('"') and formula.endswith('"'):
            values = [v.strip() for v in formula.strip('"').split(",")]
            logger.debug(
                f"[{marketplace}] DV #{dv_index}: "
                f"прямой список, {len(values)} значений"
            )
            return values

        # Именованный диапазон
        if formula in named_ranges:
            try:
                range_formula = named_ranges[formula]
                clean = range_formula.replace("$", "")
                if "!" in clean:
                    sheet_name, range_ref = clean.split("!", 1)
                    target_ws = ws.parent[sheet_name.strip("'")]
                else:
                    range_ref = clean
                    target_ws = ws

                values = [
                    str(cell.value).strip()
                    for row in target_ws[range_ref]
                    for cell in row
                    if cell.value is not None
                ]
                logger.info(
                    f"✅ [{marketplace}] DV #{dv_index}: "
                    f"{len(values)} значений из именованного диапазона '{formula}'"
                )
                return values
            except Exception as e:
                logger.error(
                    f"[{marketplace}] DV #{dv_index}: "
                    f"ошибка именованного диапазона '{formula}': {e}"
                )
                return []

        # Диапазон ячеек: Sheet1!$A$1:$A$10 или $A$1:$A$10
        if ":" in formula:
            try:
                clean = formula.replace("$", "")
                if "!" in clean:
                    sheet_name, range_ref = clean.split("!", 1)
                    target_ws = ws.parent[sheet_name.strip("'")]
                else:
                    range_ref = clean
                    target_ws = ws

                values = [
                    str(cell.value).strip()
                    for row in target_ws[range_ref]
                    for cell in row
                    if cell.value is not None
                ]
                logger.debug(
                    f"[{marketplace}] DV #{dv_index}: "
                    f"{len(values)} значений из диапазона"
                )
                return values
            except Exception as e:
                logger.error(
                    f"[{marketplace}] DV #{dv_index}: "
                    f"ошибка диапазона '{formula}': {e}"
                )
                return []

        logger.warning(
            f"[{marketplace}] DV #{dv_index}: "
            f"неизвестный формат формулы: '{formula}'"
        )
        return []

    # ------------------------------------------------------------------
    # Сохранение результатов
    # ------------------------------------------------------------------

    def _build_excel_column_mapping(
        self, ws, header_row: int
    ) -> Dict[str, List[int]]:
        """
        Строит маппинг {df_столбец: [excel_индекс_1, ...]}.

        Учитывает дубликаты: первое вхождение — оригинальное имя,
        последующие — с суффиксом («Вес (кг)1»), и оба маппируются
        на свои Excel-столбцы.

        Args:
            ws:         лист openpyxl.
            header_row: номер строки с заголовками.

        Returns:
            Словарь {df_имя_столбца: [список индексов Excel-столбцов]}.
        """
        mapping: Dict[str, List[int]] = {}
        header_count: Dict[str, int] = {}

        for col_idx, cell in enumerate(ws[header_row], start=1):
            if not cell.value:
                continue

            col_name = str(cell.value).strip()

            if col_name not in header_count:
                header_count[col_name] = 0
                df_col = col_name
            else:
                header_count[col_name] += 1
                df_col = f"{col_name}{header_count[col_name]}"

            mapping.setdefault(df_col, []).append(col_idx)

        return mapping

    def _sync_duplicate_columns(
        self, df: pd.DataFrame, marketplace: str
    ) -> None:
        """
        Копирует значения из оригинального столбца в его дубликат.

        Например, если «Вес (кг)» был переименован в «Вес (кг)1» при загрузке,
        перед записью значения из «Вес (кг)» копируются в «Вес (кг)1»,
        чтобы оба Excel-столбца содержали актуальные данные.

        Args:
            df:          DataFrame маркетплейса (изменяется на месте).
            marketplace: ключ маркетплейса.
        """
        if marketplace not in self.original_column_names:
            return

        renamed_map = self.original_column_names[marketplace]["renamed"]

        for duplicated_name, original_name in renamed_map.items():
            if original_name not in df.columns or duplicated_name not in df.columns:
                continue

            for idx in df.index:
                original_value = df.at[idx, original_name]
                if pd.notna(original_value) and str(original_value).strip():
                    df.at[idx, duplicated_name] = original_value

            logger.info(
                f"✅ [{marketplace}] Синхронизированы дубликаты: "
                f"'{original_name}' → '{duplicated_name}'"
            )

    @staticmethod
    def _expand_sheet_if_needed(
        ws, data_start_row: int, df_len: int
    ) -> None:
        """
        Расширяет лист, копируя стили последней строки для новых строк.

        Args:
            ws:            лист openpyxl.
            data_start_row: первая строка данных.
            df_len:         количество строк в DataFrame.
        """
        current_rows = ws.max_row
        required_rows = data_start_row + df_len

        if required_rows <= current_rows:
            return

        logger.info(
            f"[INFO] Расширяю лист: {current_rows} → {required_rows} строк"
        )
        last_data_row = current_rows

        for row_idx in range(current_rows + 1, required_rows + 1):
            for col_idx in range(1, ws.max_column + 1):
                source = ws.cell(row=last_data_row, column=col_idx)
                target = ws.cell(row=row_idx, column=col_idx)

                if source.has_style:
                    target.font       = source.font.copy()
                    target.border     = source.border.copy()
                    target.fill       = source.fill.copy()
                    target.number_format = source.number_format
                    target.protection = source.protection.copy()
                    target.alignment  = source.alignment.copy()

    async def _write_dataframe_to_sheet(
        self,
        ws,
        df: pd.DataFrame,
        data_start_row: int,
        df_to_excel_mapping: Dict[str, List[int]],
        marketplace: str,
        stats: Dict[str, int],
    ) -> None:
        """
        Записывает DataFrame в лист openpyxl построчно.

        Для каждой ячейки с data validation запускает AI-проверку.
        Пропускает NaN-значения. Дублирующиеся столбцы записываются
        во все соответствующие Excel-столбцы.

        Args:
            ws:                  лист openpyxl.
            df:                  DataFrame маркетплейса.
            data_start_row:      первая строка данных на листе.
            df_to_excel_mapping: маппинг {df_столбец: [excel_индексы]}.
            marketplace:         ключ МП (для логирования).
            stats:               счётчик статистики (изменяется на месте).
        """
        for row_num, (_, row) in enumerate(df.iterrows()):
            excel_row_idx = data_start_row + row_num

            for df_col_name, value in row.items():
                if pd.isna(value):
                    continue

                excel_cols = df_to_excel_mapping.get(df_col_name, [])

                for excel_col_idx in excel_cols:
                    cell = ws.cell(
                        row=excel_row_idx, column=excel_col_idx
                    )
                    allowed_values = self._get_validation_list_values(
                        ws, excel_row_idx, excel_col_idx
                    )

                    if allowed_values and self._ai_comparator:
                        matched = await self._ai_comparator.match_value_with_list(
                            str(value), allowed_values
                        )
                        if matched:
                            cell.value = matched
                            stats["ai_matched"] += 1
                        else:
                            stats["validation_conflicts"] += 1
                            stats["skipped"] += 1
                            continue
                    else:
                        cell.value = value

                    stats["saved"] += 1

    @staticmethod
    def _get_validation_list_values(
        ws, row_idx: int, col_idx: int
    ) -> List[str]:
        """
        Получает список допустимых значений для конкретной ячейки.

        Проверяет все data validation правила листа и возвращает список
        значений для правила типа «list», покрывающего указанную ячейку.

        Args:
            ws:      лист openpyxl.
            row_idx: номер строки ячейки.
            col_idx: номер столбца ячейки.

        Returns:
            Список допустимых значений или пустой список.
        """
        from openpyxl.utils import range_boundaries

        cell_ref = f"{get_column_letter(col_idx)}{row_idx}"

        for dv in ws.data_validations.dataValidation:
            if dv.type != "list" or dv.sqref is None:
                continue

            cell_in_range = False
            for range_str in str(dv.sqref).split():
                if ":" in range_str:
                    try:
                        min_col, min_row, max_col, max_row = range_boundaries(
                            range_str
                        )
                        if (
                            min_col <= col_idx <= max_col
                            and min_row <= row_idx <= max_row
                        ):
                            cell_in_range = True
                            break
                    except Exception:
                        pass
                elif range_str == cell_ref:
                    cell_in_range = True
                    break

            if not cell_in_range:
                continue

            if not dv.formula1:
                return []

            formula = dv.formula1

            if formula.startswith('"') and formula.endswith('"'):
                return [v.strip() for v in formula.strip('"').split(",")]

            if ":" in formula:
                try:
                    clean = formula.replace("$", "")
                    if "!" in clean:
                        sheet_name, range_ref = clean.split("!", 1)
                        target_ws = ws.parent[sheet_name.strip("'")]
                    else:
                        range_ref = clean
                        target_ws = ws

                    return [
                        str(cell.value).strip()
                        for row in target_ws[range_ref]
                        for cell in row
                        if cell.value is not None
                    ]
                except Exception as e:
                    logger.error(
                        f"Ошибка извлечения validation из диапазона "
                        f"'{formula}': {e}"
                    )
                    return []

        return []

    @staticmethod
    def _log_save_stats(stats: Dict[str, int]) -> None:
        """
        Выводит итоговую статистику сохранения.

        Args:
            stats: словарь со счётчиками операций сохранения.
        """
        logger.info("\n" + "=" * 60)
        logger.info("СТАТИСТИКА СОХРАНЕНИЯ:")
        logger.info(f"  ✓ Записано:        {stats['saved']}")
        logger.info(f"  🤖 AI-сопоставлений: {stats['ai_matched']}")
        logger.info(f"  ⚠ Конфликтов:      {stats['validation_conflicts']}")
        logger.info("=" * 60)

    # ------------------------------------------------------------------
    # AI-логи
    # ------------------------------------------------------------------

    def _create_ai_log_sheet(
        self, output_path: str, ai_validation_log: List[Dict]
    ) -> None:
        """
        Добавляет лист «AI_Логи» в выходной файл МП.

        Лист вставляется первым (индекс 0).

        Args:
            output_path:       путь к выходному файлу.
            ai_validation_log: список записей AI-валидации.
        """
        self._write_ai_log_sheet(
            output_path, ai_validation_log, sheet_index=0
        )

    @staticmethod
    def _write_ai_log_sheet(
        file_path: str,
        ai_validation_log: List[Dict],
        sheet_index: int,
    ) -> None:
        """
        Создаёт лист «AI_Логи» в xlsx-файле по указанному пути.

        Если лист уже существует — удаляет и пересоздаёт.
        Применяет форматирование заголовков и автоширину столбцов.

        Args:
            file_path:         путь к xlsx-файлу.
            ai_validation_log: список словарей с полями валидации.
            sheet_index:       позиция листа (0 = первый, 1 = второй).
        """
        _HEADERS = [
            "Маркетплейс",
            "Столбец",
            "Исходное значение",
            "Сопоставлено с",
            "Метод",
        ]
        _HEADER_FILL  = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        _HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
        _HEADER_ALIGN = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        try:
            wb = load_workbook(file_path)

            if "AI_Логи" in wb.sheetnames:
                del wb["AI_Логи"]

            ws = wb.create_sheet("AI_Логи", sheet_index)
            ws.freeze_panes = "A2"

            # Заголовки
            for col_idx, header in enumerate(_HEADERS, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill      = _HEADER_FILL
                cell.font      = _HEADER_FONT
                cell.alignment = _HEADER_ALIGN

            # Данные
            for row_idx, row_data in enumerate(ai_validation_log, start=2):
                for col_idx, key in enumerate(
                    ["Маркетплейс", "Столбец",
                     "Исходное значение", "Сопоставлено с", "Метод"],
                    start=1,
                ):
                    ws.cell(
                        row=row_idx, column=col_idx,
                        value=row_data.get(key, "")
                    )

            # Автоширина столбцов
            for col in ws.columns:
                max_length = max(
                    (
                        len(str(cell.value))
                        for cell in col
                        if cell.value is not None
                    ),
                    default=0,
                )
                ws.column_dimensions[col[0].column_letter].width = min(
                    max_length + 3, 50
                )

            wb.save(file_path)
            logger.info(f"✅ Лист 'AI_Логи' добавлен в {file_path}")

        except Exception as e:
            logger.error(f"❌ Ошибка при создании AI_Логи в '{file_path}': {e}")