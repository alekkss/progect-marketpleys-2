"""
Модуль для чтения данных из Excel файлов.

Отвечает за безопасное чтение столбцов из Excel файлов маркетплейсов (WB, Ozon, Яндекс).
Обрабатывает известные проблемы с data validation в файлах Ozon.

Принцип Single Responsibility: только чтение Excel, без бизнес-логики.
"""

import sys
import warnings
from pathlib import Path
from typing import List, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.cell_range import MultiCellRange

from utils.logger_config import setup_logger

sys.path.insert(0, str(Path(__file__).parent.parent))

logger = setup_logger('excel_reader')


def _patch_data_validation() -> None:
    """
    Патчит класс DataValidation в openpyxl для обработки
    некорректных sqref из файлов Ozon.

    Проблема: файлы Ozon содержат data validation с форматом sqref,
    который openpyxl не может распарсить, что вызывает TypeError.
    Решение: перехватываем ошибку и устанавливаем пустой диапазон.

    Важно: нельзя подставлять None вместо sqref — MultiCellRange(None)
    вызовет set(None), что даёт TypeError. Используем MultiCellRange()
    (пустой объект) как безопасную замену.
    """
    original_init = DataValidation.__init__

    def _safe_init(self, *args, **kwargs) -> None:
        """Безопасная инициализация DataValidation с обработкой ошибок sqref."""
        sqref = kwargs.get('sqref', None)

        # Если sqref=None пришёл из XML — сразу заменяем на пустой диапазон,
        # иначе MultiCellRange(None) упадёт с TypeError: 'NoneType' is not iterable
        if sqref is None and 'sqref' in kwargs:
            kwargs['sqref'] = MultiCellRange()
        elif sqref is not None:
            try:
                # Пробуем преобразовать sqref заранее, чтобы отловить некорректные значения
                if not isinstance(sqref, MultiCellRange):
                    MultiCellRange(sqref)
            except (TypeError, ValueError):
                logger.warning(
                    f"Некорректный sqref в data validation: {sqref!r}. "
                    f"Заменяем на пустой диапазон."
                )
                kwargs['sqref'] = MultiCellRange()

        try:
            original_init(self, *args, **kwargs)
        except TypeError as e:
            logger.warning(
                f"Ошибка инициализации DataValidation: {e}. "
                f"Создаём без sqref."
            )
            kwargs['sqref'] = MultiCellRange()
            original_init(self, *args, **kwargs)

    DataValidation.__init__ = _safe_init


# Применяем патч при импорте модуля
_patch_data_validation()


class ExcelReader:
    """Класс для чтения столбцов из Excel файлов маркетплейсов."""

    @staticmethod
    def get_column_names(file_path: str, sheet_name: str, row_number: int) -> List[str]:
        """
        Получает названия столбцов из указанной строки Excel файла.

        Безопасно обрабатывает файлы с data validation (Ozon)
        и файлы с нестандартной структурой (Яндекс.Маркет).

        Args:
            file_path: путь к файлу Excel
            sheet_name: название листа
            row_number: номер строки (1-индексация)

        Returns:
            Список названий столбцов

        Raises:
            FileNotFoundError: если файл не найден
            KeyError: если лист не найден в файле
            ValueError: если строка пустая или не содержит данных
        """
        logger.info(
            f"Чтение столбцов из файла: {Path(file_path).name}, "
            f"лист: '{sheet_name}', строка: {row_number}"
        )

        workbook = None
        try:
            # data_only=True — читаем значения, а не формулы
            # Подавляем предупреждения openpyxl о data validation
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                workbook = load_workbook(file_path, data_only=True)

            if sheet_name not in workbook.sheetnames:
                available = ', '.join(workbook.sheetnames)
                raise KeyError(
                    f"Лист '{sheet_name}' не найден. "
                    f"Доступные листы: {available}"
                )

            worksheet = workbook[sheet_name]

            # Получаем все значения из указанной строки
            row_values = []
            for cell in worksheet[row_number]:
                if cell.value is not None:
                    row_values.append(str(cell.value).strip())
                else:
                    row_values.append("")

            non_empty_count = sum(1 for v in row_values if v)
            logger.info(
                f"Прочитано столбцов: {len(row_values)}, "
                f"непустых: {non_empty_count}"
            )

            return row_values

        except KeyError:
            raise
        except TypeError as e:
            logger.error(
                f"Ошибка типа при чтении файла {Path(file_path).name}: {e}",
                exc_info=True
            )
            raise ValueError(
                f"Не удалось прочитать файл '{Path(file_path).name}'. "
                f"Возможно, файл повреждён или имеет нестандартный формат."
            ) from e
        except Exception as e:
            logger.error(
                f"Непредвиденная ошибка при чтении файла "
                f"{Path(file_path).name}: {e}",
                exc_info=True
            )
            raise
        finally:
            if workbook is not None:
                workbook.close()

    @staticmethod
    def find_column_fuzzy(columns: List[str], search_term: str) -> Optional[str]:
        """
        Ищет столбец по части названия (нечеткий поиск).

        Args:
            columns: список названий столбцов
            search_term: искомый термин

        Returns:
            Найденное название столбца или None
        """
        if not search_term:
            return None

        # Точное совпадение
        if search_term in columns:
            return search_term

        # Ищем по ключевым словам
        search_lower = search_term.lower()
        for col in columns:
            if search_lower in col.lower():
                return col

        return None
